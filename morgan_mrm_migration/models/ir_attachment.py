import base64
from datetime import timedelta,datetime
from odoo import models, fields, api
from odoo.exceptions import UserError
import xlrd
import string
from zipfile import ZipFile
from openpyxl import load_workbook
import dateparser
from dateutil.parser import parse
from odoo import SUPERUSER_ID

regions = {}
regions["AUH"] = 21
regions["ACI"] = 63
regions["AAD"] = 31
regions["UCB"] = 16
regions["DOH"] = 87
regions["ALQ"] = 88
regions["ALX"] = 89
regions["AUT"] = 18
regions["ADJ"] = 90
regions["AUS"] = 91
regions["UOB"] = 92
regions["BGL"] = 93
regions["ABG"] = 2
regions["BNGKey"] = 155
regions["BAL"] = 94
regions["BAN"] = 95
regions["ROF"] = 55
regions["BAU"] = 96
regions["BRU"] = 97
regions["CGT"] = 98
regions["AUC"] = 99
regions["CAL"] = 100
regions["CHD"] = 101
regions["CHG"] = 102
regions["CHN"] = 28
regions["CHR"] = 103
regions["CHL"] = 104
regions["CHE"] = 105
regions["COC"] = 32
regions["COL"] = 108
regions["COH"] = 109
regions["CSL"] = 110
regions["COR"] = 111
regions["DMS"] = 112
regions["TAD"] = 36
regions["DXB"] = 15
regions["EGY"] = 113
regions["LUX"] = 114
regions["GNV"] = 115
regions["HDB"] = 29
regions["HRB"] = 116
regions["HDT"] = 117
regions["TRI"] = 73
regions["HYL"] = 118
regions["IND"] = 119
regions["QAR"] = 120
regions["INTL"] = 121
regions["BHR"] = 122
regions["JAI"] = 123
regions["TAJ"] = 37
regions["JNI"] = 20
regions["JUB"] = 124
regions["KAR"] = 125
regions["KLK"] = 126
regions["KOL"] = 127
regions["KWT"] = 128
regions["KWO"] = 129
regions["LAC"] = 130
regions["BHM"] = 131
regions["BIBF"] = 132
regions["BHD"] = 133
regions["MAN"] = 11
regions["MAR"] = 134
regions["MOR"] = 135
regions["MUB"] = 136
regions["MUM"] = 137
regions["OMR"] = 138
regions["SHAH"] = 139
regions["DLH"] = 27
regions["NDEL"] = 140
regions["DEL"] = 141
regions["OL"] = 142
regions["OTH"] = 143
regions["OTW"] = 144
regions["PKS"] = 145
regions["PAR"] = 146
regions["PUN"] = 26
regions["PHO"] = 147
regions["PNQ"] = 148
regions["QFC"] = 149
regions["QBC"] = 7
regions["TAR"] = 35
regions["SAS"] = 150
regions["TER"] = 13
regions["TOR"] = 151
regions["USK"] = 152
regions["VCR"] = 153
regions["ZUR"] = 154


class Attachment(models.Model):
    _inherit = 'ir.attachment'
    processed_rows_count = fields.Integer('# of processed rows')
    is_invoice = fields.Boolean('Is invoice')
    is_processed = fields.Boolean('Processed')
    lines_processed = fields.Boolean()
    quotation_lines_row_count = fields.Integer('# of quotation lines')
    is_receipt = fields.Boolean()
    is_event = fields.Boolean()
    is_iht_event = fields.Boolean()
    is_session = fields.Boolean()
    is_iht_session = fields.Boolean()
    is_public_evaluation = fields.Boolean()
    is_iht_evaluation = fields.Boolean()



    def import_events(self):

        currencies = {
        'Bahraini Dinar': self.env.ref('base.BHD').id,
        'Canadian Dollars ': self.env.ref('base.CAD').id,
        'Indian Rupee': self.env.ref('base.INR').id,
        'Jordanian Dinar': self.env.ref('base.JOD').id,
        'Lebanese Pound': self.env.ref('base.LBP').id,
        'Saudi Riyal': self.env.ref('base.SAR').id,
        'U.S. Dollars': self.env.ref('base.USD').id,
        'United Arab Emirates Dirham': self.env.ref('base.AED').id,
        'Swiss Franc': self.env.ref('base.CHF').id,
        'Pounds Sterling': self.env.ref('base.GBP').id,
        }
        

        count = self.processed_rows_count
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        new_value = count + 500
        if new_value > len(reader):
            new_value = len(reader)
            
            self.is_processed = True
            

        for i in range(count, new_value):
            lines = reader[i].split('\t')
            vals = {}
            vals['name'] = lines[0]
            vals['mrm_id'] = lines[1]
            
            term = self.env['term'].search([('name','=',lines[3])],limit=1)
            if not term:
                term = self.env['term'].create({
                    'name': lines[2],
                    'mrm_id': lines[3],
                    })

            vals['term_id'] = term.id
            
            program = self.env['program'].search([('mrm_id','=',int(lines[5]))],limit=1)
            if not program:
                program = self.env['program'].create({
                    'name': lines[4],
                    'mrm_id': int(lines[5]),
                    })

            vals['program_id'] = program.id

            course = self.env['course'].search([('mrm_id','=', int(lines[7])),('active','in',[True,False])],limit=1)
            vals['course_id'] = course.id

            event_type = self.env['event.type'].search([('name','=',lines[23].strip())],limit=1)
            if not event_type:
                event_type = self.env['event.type'].create({
                    'name': lines[23].strip(),
                    })

            if not course.type_id:
                course.write({'type_id': event_type.id})

            vals['event_type_id'] = event_type.id

            date_begin = lines[8].replace("/","-")
            date_begin = date_begin + " 00:00:00"

            date_end = lines[9].replace("/","-")
            date_end = date_end + " 00:00:00"

            vals['date_begin'] = datetime.strptime (date_begin, "%d-%m-%Y %H:%M:%S")
            vals['date_end'] = datetime.strptime (date_end, "%d-%m-%Y %H:%M:%S")


            region = self.env['res.company'].search([('name','=',lines[16])],limit=1)
            if not region:
                region = self.env['res.company'].search([('name','ilike',lines[16])],limit=1)
                if not region:
                    region = self.env['res.company'].search([('name','=',lines[17])],limit=1)

            vals['company_id'] = region.id

            status = 'draft'
            if lines[19] == 'TRUE':
                status = 'draft'


            if lines[20] == 'TRUE':
                vals['is_online'] = True

            if lines[21] == 'TRUE':
                status = 'cancel'

            if lines[19] == 'FALSE' and lines[21] == 'FALSE':
                status = 'confirm'

            if vals['date_begin'] < datetime.today() and lines[21] == 'FALSE':
                status = 'done'

            vals['state'] = status


            location = self.env['res.partner'].sudo().search([('name','=',lines[18].strip())])
            if not location:
                location = self.env['res.partner'].create({
                    'name': lines[18].strip(),
                    'type': 'training',
                    'company_id': region.id,
                    })

            else:
                location.write({
                    'type': 'training',
                    'company_id': region.id,
                    })

            vals['address_id'] = location.id

            class_name = 'Class A - ' + lines[18].strip()
            class_room = self.env['event.track.location'].search([('name','=',class_name),('location_id','=',location.id)])
            if not class_room:
                class_room = self.env['event.track.location'].create({
                    'name': class_name,
                    'location_id': location.id,
                    'region_id': region.id,
                    })

            vals['location_id'] = class_room.id


            event = self.env['event.event'].search([('mrm_id','=',vals['mrm_id'])],limit=1)
            if not event:
                event = self.env['event.event'].create(vals)

            else:
                event.write(vals)

            event_ticket_vals = {}
            event_ticket_vals['event_id'] = event.id
            event_ticket_vals['name'] = 'Tuition fees'
            event_ticket_vals['ticket_company_id'] = event.company_id.id
            event_ticket_vals['price'] = float(lines[24])


            event.import_prices(event.mrm_id)

            #raise UserError(lines[23])
            if lines[26]:
                tax = self.env['account.tax'].search([('name','=',lines[26])])
                if not tax:
                    tax = self.env['account.tax'].search([('name','ilike',lines[26])])


                if len(tax) > 1:
                    if tax.filtered(lambda l: l.company_id.id == region.id):
                        tax = tax.filtered(lambda l: l.company_id.id == region.id)
                        

                if len(tax) > 1:
                    if tax.filtered(lambda l: l.amount == float(lines[27])):
                        tax.filtered(lambda l: l.amount == float(lines[27]))


                if not tax:
                    tax = self.env['account.tax'].search([('amount','=',float(lines[27])),('company_id','=',region.id)],limit=1)
                if event.event_ticket_ids:
                    if len(tax) > 1:
                        tax = tax[0]

                    for ticket in event.event_ticket_ids:
                        ticket.write({'tax_ids': [(6, 0, [tax.id])]})


            event.add_prices_to_course()
        
        self.write({'processed_rows_count': new_value})

    def fix_events_date(self):
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        for row in reader:
            lines = row.split('\t')
            vals = {}

            date_begin = lines[8].replace("/","-")
            date_begin = date_begin + " 00:00:00"
            date_end = lines[9].replace("/","-")
            date_end = date_end + " 00:00:00"
            vals['date_begin'] = datetime.strptime (date_begin, "%d-%m-%Y %H:%M:%S") + timedelta(hours=-2)
            vals['date_end'] = datetime.strptime (date_end, "%d-%m-%Y %H:%M:%S") + timedelta(hours=-2)


            event = self.env['event.event'].search([('mrm_id','=',vals['mrm_id'])],limit=1)
            event.write(vals)




    def import_sessions(self):
        count = self.processed_rows_count
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        new_value = count + 1000
        if new_value > len(reader):
            new_value = len(reader)
            


            self.is_processed = True
            

        for i in range(count, new_value):
            lines = reader[i].split('\t')
            vals = {}
            if lines[3]:
                vals['event_id'] = self.env['event.event'].search([('mrm_id','=',int(lines[3]))]).id
            
            vals['mrm_id'] = int(lines[5])
            vals['name'] = lines[4]
            if lines[6]:
                session_date = lines[6].replace("/","-")
                session_date = datetime.strptime (session_date, "%m-%d-%Y")
                if datetime.today() > session_date:
                    vals['stage_id'] = self.env.ref('website_event_track.event_track_stage1').id


                vals['date'] = session_date

            if lines[7]:
                track_from = lines[7][:-3]
                #raise UserError(track_from)
                track_from = track_from.replace(":", ".")
                track_from_array = track_from.split(".")
                if track_from_array[1] == '3':
                    track_from = float(track_from) + 0.2

                vals['track_from'] = track_from

            if lines[8]:
                to = lines[8][:-3]
                to = to.replace(":", ".")
                track_to_array = to.split(".")
                if track_to_array[1] == '3':
                    to = float(to) + 0.2


                vals['to'] = to

            instructor = False
            if lines[10]:
                instructor = self.env['hr.employee'].search([('mrm_id','=',int(lines[10])),('active','in',[True,False])])
                if not instructor:
                    instructor = self.env['hr.employee'].create({
                        'name': lines[9],
                        'mrm_id': lines[10],
                        'is_instructor': True
                        })
                instructor = instructor.id

            vals['employee_id'] = instructor

            session = self.env['event.track'].search([('mrm_id','=',vals['mrm_id'])])
            if not session:
                if 'event_id' in vals and vals['event_id']:
                    session = self.env['event.track'].create(vals)

            else:
                session.write(vals)

        self.write({'processed_rows_count': new_value})

    def import_instructors(self):
        attachment = self.env['ir.attachment'].browse(self.id)
        objFile = str(base64.b64decode(attachment.datas).decode('UTF-8'))
        reader = objFile.splitlines()
        companies = []
        for row in reader:
            lines = row.split('\t')
            vals = {}
            vals['mrm_id'] = int(lines[0])
            vals['name'] = lines[1]

            if lines[2]:
                company_name = lines[2].rstrip()
                company = self.env['res.company'].search([('name','=',company_name)])
                if not company:
                    company = self.env['res.company'].search([('mrm_id','=',int(lines[3]))])
                vals['company_id'] = company.id

            else:
                vals['company_id'] = False
            
            vals['active'] = False if lines[4] == 'TRUE' else True
            vals['mrm_status'] = 'dormant' if lines[4] == 'TRUE' else 'active'
            vals['is_instructor'] = True

            instructor = self.env['hr.employee'].search([('mrm_id','=',vals['mrm_id']),('active','in',[True,False])])
            if not instructor:
                instructor = self.env['hr.employee'].create(vals)

            else:
                instructor = instructor.write(vals)



    def import_europe_events(self):

        currencies = {
        'Bahraini Dinar': self.env.ref('base.BHD').id,
        'Canadian Dollars ': self.env.ref('base.CAD').id,
        'Indian Rupee': self.env.ref('base.INR').id,
        'Jordanian Dinar': self.env.ref('base.JOD').id,
        'Lebanese Pound': self.env.ref('base.LBP').id,
        'Saudi Riyal': self.env.ref('base.SAR').id,
        'U.S. Dollars': self.env.ref('base.USD').id,
        'United Arab Emirates Dirham': self.env.ref('base.AED').id,
        'Euros': self.env.ref('base.EUR').id,
        'Pounds Sterling': self.env.ref('base.GBP').id,
        'Swiss Franc': self.env.ref('base.CHF').id
        }
        

        attachment = self.env['ir.attachment'].browse(self.id)
        objFile = str(base64.b64decode(attachment.datas).decode('UTF-8'))
        reader = objFile.splitlines()
        for row in reader:
            lines = row.split('\t')
            vals = {}
            vals['tfrm_id'] = lines[0]
            
            if lines[3]:
                term = self.env['term'].search([('name','=',lines[2])])
                if not term:
                    term = self.env['term'].create({
                        'name': lines[2],
                        'mrm_id': lines[3],
                        })

                else:
                    term.write({'mrm_id': lines[3]})

                vals['term_id'] = term.id
            
            if lines[5]:
                if lines[5] != '10, 11, 12, 6, 7, 8, 9':
                    program = self.env['program'].search([('mrm_id','=',int(lines[5]))],limit=1)
                    if not program:
                        program = self.env['program'].create({
                            'name': lines[4],
                            'mrm_id': int(lines[5]),
                            })

                    vals['program_id'] = program.id

            course = False
            if lines[8]:
                course = self.env['course'].search([('tfrm_id','=', int(lines[8]))])

            if not course:
                course = self.env['course'].search([('name','=', lines[7])])

            if not course and lines[8] and lines[7] and lines[5] and lines[5] != '10, 11, 12, 6, 7, 8, 9':
                course = self.env['course'].create({
                    'name': lines[7],
                    'tfrm_id': lines[8],
                    'program_id': vals['program_id']
                    })

            if course:
                vals['course_id'] = course.id


            event_type = self.env['event.type'].search([('name','=',lines[23])],limit=1)
            if not event_type:
                event_type = self.env['event.type'].create({
                    'name': lines[23],
                    })

            if not course.type_id:
                course.write({'type_id': event_type.id})

            vals['event_type_id'] = event_type.id

            vals['name'] = lines[1]

            date_begin = lines[9].replace("/","-")
            date_begin = date_begin[:-2] + "20" + date_begin[-2:]
            date_begin = date_begin + " 00:00:00"

            date_end = lines[10].replace("/","-")
            date_end = date_end[:-2] + "20" + date_end[-2:]
            date_end = date_end + " 00:00:00"

            vals['date_begin'] = datetime.strptime (date_begin, "%d-%m-%Y %H:%M:%S")
            vals['date_end'] = datetime.strptime (date_end, "%d-%m-%Y %H:%M:%S")


            if lines[16]:
                region = self.env['res.company'].search([('name','=',lines[16].strip())],limit=1)
                if not region:
                    region = self.env['res.company'].search([('name','ilike',lines[16].strip())],limit=1)
                vals['company_id'] = region.id

            status = 'draft'
            if lines[20] == 'TRUE':
                status = 'draft'


            if lines[21] == 'TRUE':
                vals['is_online'] = True

            if lines[22] == 'TRUE':
                status = 'cancel'

            vals['state'] = status


            location = self.env['res.partner'].search([('type','=','training'),('name','=',lines[19])])
            if not location:
                location = self.env['res.partner'].create({
                    'name': lines[19],
                    'type': 'training',
                    'company_id': region.id,
                    })

            vals['address_id'] = location.id

            class_name = 'Class A - ' + lines[19]
            class_room = self.env['event.track.location'].search([('name','=',class_name),('location_id','=',location.id)])
            if not class_room:
                class_room = self.env['event.track.location'].create({
                    'name': class_name,
                    'location_id': location.id,
                    'region_id': region.id,
                    })

            vals['location_id'] = class_room.id


            event = self.env['event.event'].search([('tfrm_id','=',vals['tfrm_id'])],limit=1)
            if not event:
                event = self.env['event.event'].create(vals)

            else:
                event.write(vals)

            event_ticket_vals = {}
            event_ticket_vals['event_id'] = event.id
            event_ticket_vals['name'] = 'Tuition fees'
            event_ticket_vals['ticket_company_id'] = event.company_id.id
            event_ticket_vals['price'] = float(lines[24])


            event.import_europe_prices(event.mrm_id)

            #raise UserError(lines[23])
            if lines[26]:
                tax = self.env['account.tax'].search([('name','=',lines[26])])
                if not tax:
                    raise UserError(lines[26])
                if event.event_ticket_ids:
                    for ticket in event.event_ticket_ids:
                        ticket.write({'tax_ids': [(6, 0, [tax.id])]})

    def import_europe_iht_events(self):

        currencies = {
        'Bahraini Dinar': self.env.ref('base.BHD').id,
        'Canadian Dollars ': self.env.ref('base.CAD').id,
        'Indian Rupee': self.env.ref('base.INR').id,
        'Jordanian Dinar': self.env.ref('base.JOD').id,
        'Lebanese Pound': self.env.ref('base.LBP').id,
        'Saudi Riyal': self.env.ref('base.SAR').id,
        'U.S. Dollars': self.env.ref('base.USD').id,
        'United Arab Emirates Dirham': self.env.ref('base.AED').id,
        'Euros': self.env.ref('base.EUR').id,
        'Pounds Sterling': self.env.ref('base.GBP').id,
        'Swiss Franc': self.env.ref('base.CHF').id
        }
        

        attachment = self.env['ir.attachment'].browse(self.id)
        objFile = str(base64.b64decode(attachment.datas).decode('UTF-8'))
        reader = objFile.splitlines()
        for row in reader:
            lines = row.split('\t')
            vals = {}

            vals['name'] = lines[0]
            vals['tfrm_id'] = lines[1]
            course = False
            
            if lines[3]:
                term = self.env['term'].search([('name','=',lines[2])])
                if not term:
                    term = self.env['term'].create({
                        'name': lines[2],
                        'mrm_id': lines[3],
                        })

                else:
                    term.write({'mrm_id': lines[3]})

                vals['term_id'] = term.id


            if lines[5]:
                course = self.env['course'].search([('tfrm_id','=', int(lines[5]))])
                vals['course_id'] = course.id


            date_begin = lines[6].replace("/","-")
            date_begin = date_begin[:-2] + "20" + date_begin[-2:]
            date_begin = date_begin + " 00:00:00"

            date_end = lines[7].replace("/","-")
            date_end = date_end[:-2] + "20" + date_end[-2:]
            date_end = date_end + " 00:00:00"

            event_type = self.env['event.type'].search([('name','=',lines[19])],limit=1)
            if not event_type:
                event_type = self.env['event.type'].create({
                    'name': lines[19],
                    })

            if course and not course.type_id:
                course.write({'type_id': event_type.id})

            vals['event_type_id'] = event_type.id
            vals['date_begin'] = datetime.strptime (date_begin, "%d-%m-%Y %H:%M:%S")
            vals['date_end'] = datetime.strptime (date_end, "%d-%m-%Y %H:%M:%S")


            region = self.env['res.company'].search([('name','=',lines[12])],limit=1)
            if not region:
                region = self.env['res.company'].search([('name','ilike',lines[12])],limit=1)
            
            vals['company_id'] = region.id

            status = 'draft'
            if lines[16] == 'TRUE':
                status = 'draft'


            if lines[17] == 'TRUE':
                vals['is_online'] = True

            if lines[18] == 'TRUE':
                status = 'cancel'

            vals['state'] = status


            location = self.env['res.partner'].search([('type','=','training'),('name','=',lines[15])])
            if not location:
                location = self.env['res.partner'].create({
                    'name': lines[15],
                    'type': 'training',
                    'company_id': region.id,
                    })

            vals['address_id'] = location.id

            class_name = 'Class A - ' + lines[15]
            class_room = self.env['event.track.location'].search([('name','=',class_name),('location_id','=',location.id)])
            if not class_room:
                class_room = self.env['event.track.location'].create({
                    'name': class_name,
                    'location_id': location.id,
                    'region_id': region.id,
                    })

            vals['location_id'] = class_room.id

            partner = self.env['res.partner'].search([('tfrm_id','=',int(lines[24]))])
            if not partner:
                partner = self.env['res.partner'].create({
                    'name': lines[23],
                    'tfrm_id': int(lines[24]),
                    'from_excel': True,
                    })
            else:
                partner.write({
                    'name': lines[23],
                    'tfrm_id': int(lines[24]),
                    'from_excel': True,
                    })


            vals['partner_id'] = partner.id



            event = self.env['event.event'].search([('tfrm_id','=',vals['tfrm_id'])],limit=1)
            if not event:
                event = self.env['event.event'].create(vals)

            else:
                event.write(vals)

            if lines[20]:
                event_ticket_vals = {}
                event_ticket_vals['event_id'] = event.id
                event_ticket_vals['name'] = 'Tuition fees'
                event_ticket_vals['ticket_company_id'] = event.company_id.id
                event_ticket_vals['price'] = float(lines[20])


                event.import_europe_prices(event.tfrm_id)

                #raise UserError(lines[23])
                if lines[26]:
                    tax = self.env['account.tax'].search([('name','=',lines[21])])
                    if not tax:
                        raise UserError(lines[21])
                    if event.event_ticket_ids:
                        for ticket in event.event_ticket_ids:
                            ticket.write({'tax_ids': [(6, 0, [tax.id])]})


            #event.add_prices_to_course()

    def import_europe_instructors(self):
        attachment = self.env['ir.attachment'].browse(self.id)
        objFile = str(base64.b64decode(attachment.datas).decode('UTF-8'))
        reader = objFile.splitlines()
        companies = []
        for row in reader:
            lines = row.split('\t')
            vals = {}
            vals['tfrm_id'] = int(lines[1])
            vals['name'] = lines[0]

            if lines[2]:
                company_name = lines[2].rstrip()
                company = self.env['res.company'].search([('name','=','Morgan Europe SAS «Top Finance»')])
                vals['company_id'] = company.id

            else:
                vals['company_id'] = False

            vals['private_email'] = lines[4]
            
            vals['active'] = False if lines[9] == 'TRUE' else True
            vals['mrm_status'] = 'dormant' if lines[9] == 'TRUE' else 'active'
            vals['is_instructor'] = True
            vals['mobile_phone'] = lines[10]

            instructor = self.env['hr.employee'].search([('tfrm_id','=',vals['tfrm_id'])])
            if not instructor:
                instructor = self.env['hr.employee'].create(vals)

            else:
                instructor = instructor.write(vals)


    def import_europe_sessions(self):
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()
        count = self.processed_rows_count
        new_value = count + 1000
        if new_value > len(reader):
            new_value = len(reader)

        for i in range(count, new_value):
            lines = reader[i].split('\t')
            vals = {}
            if lines[3]:
                vals['event_id'] = self.env['event.event'].search([('tfrm_id','=',int(lines[3]))]).id
            
            
            try:
                vals['tfrm_id'] = int(lines[5])
            except:
                raise UserError(str(lines) + '   ' + str(i))
            vals['name'] = lines[4]
            session_date = lines[6].replace("/","-")
            #session_date = session_date[:-2] + "20" + session_date[-2:]
            session_date = datetime.strptime (session_date, "%d-%m-%Y")
            if datetime.today() > session_date:
                vals['stage_id'] = self.env.ref('website_event_track.event_track_stage1').id


            vals['date'] = session_date

            track_from = lines[8].replace(":", ".")
            track_from_array = track_from.split(".")
            if track_from_array[1] == '30':
                track_from = float(track_from) + 0.2

            vals['track_from'] = track_from

            to = lines[9].replace(":", ".")
            track_to_array = to.split(".")
            if track_to_array[1] == '30':
                to = float(to) + 0.2


            vals['to'] = to


            instructor = False
            if lines[15]:
                instructor = self.env['hr.employee'].search([('tfrm_id','=',int(lines[15])),('active','in',[True,False])])
                if not instructor:
                    instructor = self.env['hr.employee'].create({
                        'name': lines[14],
                        'tfrm_id': lines[15],
                        'is_instructor': True
                        })
                instructor = instructor.id

            vals['employee_id'] = instructor

            session = self.env['event.track'].search([('tfrm_id','=',vals['tfrm_id'])])
            if not session:
                if 'event_id' in vals and vals['event_id']:
                    session = self.env['event.track'].create(vals)

            else:
                session.write(vals)

        self.write({'processed_rows_count': new_value})

    def import_europe_iht_sessions(self):
        attachment = self.env['ir.attachment'].browse(self.id)
        objFile = str(base64.b64decode(attachment.datas).decode('UTF-8'))
        reader = objFile.splitlines()
        for row in reader:
            lines = row.split('\t')
            vals = {}           
            if lines[5]:
                vals['event_id'] = self.env['event.event'].search([('tfrm_id','=',int(lines[5]))]).id

            vals['tfrm_id'] = int(lines[6])
            vals['name'] = lines[7]

            
            if lines[8]:
                session_date = lines[8].replace("/","-")
                session_date = session_date[:-2] + "20" + session_date[-2:]
                session_date = datetime.strptime (session_date, "%d-%m-%Y")
                if datetime.today() > session_date:
                    vals['stage_id'] = self.env.ref('website_event_track.event_track_stage1').id


                vals['date'] = session_date

            track_from = lines[10].replace(":", ".")
            track_from_array = track_from.split(".")
            if track_from_array[1] == '30':
                track_from = float(track_from) + 0.2

            elif track_from_array[1] == '15':
                track_from = float(track_from) + 0.2

            elif track_from_array[1] == '45':
                track_from = float(track_from) + 0.2

            vals['track_from'] = track_from

            to = lines[11].replace(":", ".")
            track_to_array = to.split(".")
            if track_to_array[1] == '30':
                to = float(to) + 0.2

            elif track_to_array[1] == '15':
                to = float(to) + 0.1

            elif track_to_array[1] == '45':
                to = float(to) + 0.3


            vals['to'] = to

            instructor = False
            if lines[15]:
                instructor = self.env['hr.employee'].search([('tfrm_id','=',int(lines[15])),('active','in',[True,False])])
                if not instructor:
                    instructor = self.env['hr.employee'].create({
                        'name': lines[14],
                        'tfrm_id': lines[15],
                        'is_instructor': True
                        })
                instructor = instructor.id

            vals['employee_id'] = instructor

            session = self.env['event.track'].search([('tfrm_id','=',vals['tfrm_id'])])
            if not session:
                if 'event_id' in vals and vals['event_id']:
                    session = self.env['event.track'].create(vals)

            else:
                session.write(vals)


    def unzip_file(self):
        with ZipFile(self._full_path(self.store_fname), 'r') as zip_ref:
            zip_ref.extractall()


    def import_iht_events(self):

        currencies = {
        'Bahraini Dinar': self.env.ref('base.BHD').id,
        'Canadian Dollars ': self.env.ref('base.CAD').id,
        'Indian Rupee': self.env.ref('base.INR').id,
        'Jordanian Dinar': self.env.ref('base.JOD').id,
        'Lebanese Pound': self.env.ref('base.LBP').id,
        'Saudi Riyal': self.env.ref('base.SAR').id,
        'U.S. Dollars': self.env.ref('base.USD').id,
        'United Arab Emirates Dirham': self.env.ref('base.AED').id,
        }

        attachment = self.env['ir.attachment'].browse(self.id)
        objFile = str(base64.b64decode(attachment.datas).decode('UTF-8'))
        reader = objFile.splitlines()
        for row in reader:
            lines = row.split('\t')
            vals = {}
            vals['mrm_id'] = lines[1]
            
            term = self.env['term'].search([('name','=',lines[2].strip()),('mrm_id','=',int(lines[3]))],limit=1)
            if not term:
                term = self.env['term'].create({
                    'name': lines[2].strip(),
                    'mrm_id': int(lines[3]),
                    })

            vals['term_id'] = term.id

            vals['code'] = lines[4]

            event_type = self.env['event.type'].search([('name','=',lines[20].strip())])
            if not event_type:
                event_type = self.env['event.type'].create({
                    'name': lines[20].strip(),
                    'study_format': 'inhouse',
                    })
            vals['event_type_id'] = event_type.id

            vals['name'] = lines[0]

            date_begin = lines[5].replace("/","-")
            date_begin = date_begin + " 00:00:00"

            date_end = lines[6].replace("/","-")
            date_end = date_end + " 00:00:00"

            #raise UserError(date_begin)
            vals['date_begin'] = datetime.strptime (date_begin, "%d-%m-%Y %H:%M:%S")
            vals['date_end'] = datetime.strptime (date_end, "%d-%m-%Y %H:%M:%S")


            region = self.env['res.company'].search([('name','=',lines[12])],limit=1)
            if not region:
                region = self.env['res.company'].search([('name','ilike',lines[12])],limit=1)
                if not region:
                    region = self.env['res.company'].search([('name','=',lines[14])],limit=1)


            vals['company_id'] = region.id

            status = 'draft'
            if lines[16] == 'TRUE':
                status = 'draft'


            if lines[17] == 'TRUE':
                vals['is_online'] = True

            if lines[18] == 'TRUE':
                status = 'cancel'

            if lines[16] == 'FALSE' and lines[18] == 'FALSE':
                status = 'confirm'

            if vals['date_begin'] < datetime.today() and lines[18] == 'FALSE':
                status = 'done'

            vals['state'] = status


            location = self.env['res.partner'].sudo().search([('type','=','training'),('name','=',lines[15].strip())])
            if not location:
                location = self.env['res.partner'].sudo().create({
                    'name': lines[15].strip(),
                    'type': 'training',
                    'company_id': region.id,
                    })

            vals['address_id'] = location.id

            class_name = 'Class A - ' + lines[15]
            class_room = self.env['event.track.location'].search([('name','=',class_name),('location_id','=',location.id)])
            if not class_room:
                class_room = self.env['event.track.location'].create({
                    'name': class_name,
                    'location_id': location.id,
                    'region_id': region.id,
                    })

            vals['location_id'] = class_room.id

            if lines[26]:
                partner = self.env['res.partner'].sudo().search([('mrm_id','=',int(lines[26])),('is_mrm_contact','=',False)])
                if not partner:
                    partner = self.env['res.partner'].sudo().create({
                        'name': lines[25],
                        'mrm_id': int(lines[26]),
                        'from_excel': True,
                        })
                else:
                    partner.write({'name': lines[25], 'from_excel':True})

                vals['partner_id'] = partner.id

            event = self.env['event.event'].search([('mrm_id','=',vals['mrm_id'])],limit=1)
            if not event:
                event = self.env['event.event'].create(vals)

            else:
                event.write(vals)


            event.import_prices(event.mrm_id)

            #raise UserError(lines[23])
            if lines[23]:
                tax = self.env['account.tax'].search([('name','=',lines[23])])
                if not tax:
                    tax = self.env['account.tax'].search([('amount','=',float(lines[24])),('company_id','=',region.id)],limit=1)
                if event.event_ticket_ids:
                    for ticket in event.event_ticket_ids:
                        ticket.write({'tax_ids': [(6, 0, [tax.id])]})

        self.is_processed = True


    def import_iht_sessions(self):
        attachment = self.env['ir.attachment'].browse(self.id)
        objFile = str(base64.b64decode(attachment.datas).decode('UTF-8'))
        reader = objFile.splitlines()
        for row in reader:
            lines = row.split('\t')
            vals = {}
            if lines[3]:
                vals['event_id'] = self.env['event.event'].search([('mrm_id','=',int(lines[3]))]).id
                vals['mrm_id'] = int(lines[5])
                vals['name'] = lines[4]

                if lines[6]:
                    session_date = lines[6].replace("/","-")
                    session_date = datetime.strptime (session_date, "%d-%m-%Y") + timedelta(hours=-3)
                    if datetime.today() > session_date:
                        vals['stage_id'] = self.env.ref('website_event_track.event_track_stage1').id


                    vals['date'] = session_date

                track_from = lines[7].replace(":", ".")
                track_from_array = track_from.split(".")
                if track_from_array[1] == '30':
                    track_from = float(track_from) + 0.2

                vals['track_from'] = track_from

                to = lines[8].replace(":", ".")
                track_to_array = to.split(".")
                if track_to_array[1] == '30':
                    to = float(to) + 0.2


                vals['to'] = to

                instructor = False
                if lines[10]:
                    instructor = self.env['hr.employee'].search([('mrm_id','=',int(lines[10])),('active','in',[True,False])])
                    if not instructor:
                        instructor = self.env['hr.employee'].create({
                            'name': lines[9],
                            'mrm_id': lines[10],
                            'is_instructor': True
                            })

                    instructor = instructor.id
                
                vals['employee_id'] = instructor

                session = self.env['event.track'].search([('mrm_id','=',vals['mrm_id'])])
                if not session:
                    try:
                        session = self.env['event.track'].create(vals)
                    except:
                        raise UserError(lines[0])

                else:
                    session.write(vals)

        self.is_processed = True


    def import_registrations(self):
        value = self.processed_rows_count
        attachment = self.env['ir.attachment'].browse(self.id)
        objFile = str(base64.b64decode(attachment.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        new_value = int(value) + 100
        if new_value > len(reader):
            new_value = len(reader)

            self.is_processed = True

        for i in range(int(value), new_value):
            lines = reader[i].split('\t')
            vals = {}
            vals['from_mrm'] = 1
            if lines[0]:
                partner = self.env['res.partner'].search([('mrm_id','=',int(lines[0])),('is_mrm_contact','=',True)],limit=1)
                if not partner:
                    partner = self.env['res.partner'].create({
                        'mrm_id': lines[0],
                        'name': lines[1],
                        'email': lines[2],
                        'mobile': lines[3],
                        'phone': lines[4],
                        'company_type': 'person',
                        'from_excel': True,
                        'is_mrm_contact': True,
                        })

                else:
                    partner.write({
                        'mrm_id': lines[0],
                        'name': lines[1],
                        'mobile': lines[3],
                        'phone': lines[4],
                        'email': lines[2],
                        'company_type': 'person',
                        'from_excel': True,
                        })
                
                vals['partner_id'] = partner.id

            else:
                vals['partner_id'] = 115122
            
            vals['registration_number'] = lines[5]


            if lines[6]:
                vals['state'] = 'cancel'

            else:
                vals['state'] = 'draft'
            
            event = self.env['event.event'].search([('mrm_id','=',int(lines[7]))])
            #event = self.env['event.event'].browse(int(lines[18]))
            #vals['session_id'] = session.id
            vals['event_id'] = event.id

            vals['progress'] = float(lines[10]) if lines[10] else 0


            vals['score'] = float(lines[11]) if lines[11] else 0
            
            if lines[12] == 'on':
                vals['passing_status'] = 'passed'

            if lines[13]  == 'on':
                vals['passing_status'] = 'failed'

            if not lines[12] and not lines[13]:
                vals['passing_status'] = False

            vals['reason_if_failed'] = lines[14]
            vals['mrm_id'] = lines[17]

            if 'partner_id' in vals:
                #registration = self.env['event.registration'].search([('partner_id','=',vals['partner_id']),('event_id','=',vals['event_id']),('registration_number','=',vals['registration_number'])])
                registration = self.env['event.registration'].search([('mrm_id', '=', vals['mrm_id'])])
                if not registration:
                    registration = self.env['event.registration'].create(vals)

                else:
                    registration.write(vals)

                registration._onchange_partner()

                
                """if session:
                    if lines[7] == '0' or lines[7] == '1':
                        attendance = self.env['attendance'].search([('session_id','=',session.id),('registration_id','=',registration.id)])
                        if not attendance:
                            attendance = self.env['attendance'].create({
                                'registration_id': registration.id,
                                'event_id': session.event_id.id,
                                'session_id': session.id,
                                'partner_id': vals['partner_id'],
                                'attended': True if lines[7] == '1' else False
                                })"""
        self.write({'processed_rows_count': new_value})

    def import_attendances(self):
        value = self.processed_rows_count
        attachment = self.env['ir.attachment'].browse(self.id)
        objFile = str(base64.b64decode(attachment.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        new_value = int(value) + 1000
        if new_value > len(reader):
            new_value = len(reader)

            self.is_processed = True


        for i in range(int(value), new_value):
            lines = reader[i].split('\t')
            vals = {}


            if lines[0]:
                partner = self.env['res.partner'].search([('mrm_id','=',int(lines[0])),('registration_ids','!=',False)],limit=1)
                if not partner:
                    partner = self.env['res.partner'].create({
                        'mrm_id': lines[0],
                        'name': lines[1],
                        'email': lines[2],
                        'mobile': lines[3],
                        'phone': lines[4],
                        'company_type': 'person',
                        'from_excel': True,
                        })

                else:
                    partner.write({
                        'mrm_id': lines[0],
                        'name': lines[1],
                        'mobile': lines[3],
                        'phone': lines[4],
                        'company_type': 'person',
                        'from_excel': True,
                        })
                
                vals['partner_id'] = partner.id
            
            vals['registration_number'] = lines[5]
            
            session = self.env['event.track'].search([('mrm_id','=',int(lines[7].strip()))])
            if session:
                if 'partner_id' in vals:
                    registration = self.env['event.registration'].search([('partner_id','=',vals['partner_id']),('event_id','=',session.event_id.id),('registration_number','=',vals['registration_number'])],limit=1)
                    if registration:
                        attendance = self.env['attendance'].search([('session_id','=',session.id),('registration_id','=',registration.id)])
                        if not attendance:
                            attendance = self.env['attendance'].create({
                                'registration_id': registration.id,
                                'event_id': session.event_id.id,
                                'session_id': session.id,
                                'partner_id': vals['partner_id'],
                                'attended': True if lines[10] == '1' else False
                                })
        
        self.write({'processed_rows_count': new_value})

    def import_spring_attendances(self):
        value = self.processed_rows_count
        attachment = self.env['ir.attachment'].browse(self.id)
        objFile = str(base64.b64decode(attachment.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        new_value = int(value) + 1000
        if new_value > len(reader):
            new_value = len(reader)

            self.is_processed = True

        for i in range(int(value), new_value):
            lines = reader[i].split('\t')
            vals = {}



            if lines[0]:
                partner = self.env['res.partner'].search([('mrm_id','=',int(lines[0])),('registration_ids','!=',False)],limit=1)
                vals['partner_id'] = partner.id
            
            vals['registration_number'] = lines[2]
            
            session = self.env['event.track'].search([('mrm_id','=',int(lines[3].strip()))])
            if session:
                if 'partner_id' in vals:
                    registration = self.env['event.registration'].search([('partner_id','=',vals['partner_id']),('event_id','=',session.event_id.id),('registration_number','=',vals['registration_number'])],limit=1)
                    if registration:
                        attendance = self.env['attendance'].search([('session_id','=',session.id),('registration_id','=',registration.id)])
                        if not attendance:
                            attendance = self.env['attendance'].create({
                                'registration_id': registration.id,
                                'event_id': session.event_id.id,
                                'session_id': session.id,
                                'partner_id': vals['partner_id'],
                                'attended': True if lines[5] == '1' else False,
                                'mrm_id': lines[6],
                                })
        
        self.write({'processed_rows_count': new_value})



    def import_public_evaluations(self):
        wb_obj = load_workbook(filename = 'data/filestore/morgan13-production-1206400/d6/' + self.name)
        wsheet = wb_obj['Sheet1']
        dataDict = {}

        counter = 0
        for key, *values in wsheet.iter_rows():
            key = key.value + str(counter) if key.value else str(counter)
            dataDict[key] = [v.value for v in values]
            counter = counter + 1


        for key, value in dataDict.items():
            vals = {}

            vals['name'] = value[0]
            date = value[2]
            try:
                date = date.replace("/", "-")
                date = datetime.strptime(date, "%d-%m-%Y")


            except:
                date = date

            vals['type'] = 'public'
            vals['date'] = date
            if value[3]:
                vals['average_score'] = float(value[3])

            try:
                vals['active_students'] = int(value[6])
            except:
                raise UserError(str(value))
            vals['mrm_id'] = int(value[7])

            filled_evaluation = value[8]
            if filled_evaluation == '#DIV/0!':
                filled_evaluation = 0
            else:
                filled_evaluation = float(filled_evaluation)

            vals['filled_evaluation'] = filled_evaluation

            evaluation_form = value[10]
            if evaluation_form:
                evaluation_form = evaluation_form.replace("\n","<br/>")
            vals['evaluation_form'] = evaluation_form

            session = self.env['event.track'].search([('mrm_id','=', int(value[13]))])  
            vals['session_id'] = session.id
            if session:
                vals['term_id'] = session.event_id.term_id.id
                vals['instructor_id'] = session.employee_id.id
                vals['course_id'] = session.event_id.course_id.id
                vals['region_id'] = session.event_id.company_id.id

            else:
                vals['term_id'] = self.env['term'].search([('name','=',value[4])],limit=1).id
                vals['instructor_id'] = self.env['hr.employee'].search([('mrm_id','=',int(value[12])),('active','in',[True,False])]).id
                vals['course_id'] = self.env['course'].search([('name','=',value[0]),('mrm_id','>',0)],limit=1).id

                region = self.env['res.company'].search([('name','=',value[9])])
                if not region:
                    region = self.env['res.company'].search([('name','ilike',value[9])])

                vals['region_id'] = region.id

            evaluation = self.env['mrm.instructor.evaluation'].search([('mrm_id','=',vals['mrm_id']),('session_id','=',vals['session_id'])])

            if not evaluation:                
                evaluation = self.env['mrm.instructor.evaluation'].create(vals)

            else:
                evaluation.write(vals)

        self.is_processed = True


    def import_europe_evaluations(self):
        wb_obj = load_workbook(filename = 'data/filestore/morgan13-production-1206400/d6/Instructor List for Odoo Migration.xlsx')
        wsheet = wb_obj['Instructor List for Odoo Migrat']
        dataDict = {}

        counter = 0
        for key, *values in wsheet.iter_rows():
            key = key.value + str(counter) if key.value else str(counter)
            dataDict[key] = [v.value for v in values]
            counter = counter + 1


        for key, value in dataDict.items():
            vals = {}

            vals['name'] = value[0]
            if value[1]:
                vals['instructor_id'] = self.env['hr.employee'].search([('mrm_id','=',value[1])]).id
            vals['region_id'] = self.env['res.company'].search([('name','=','Morgan Europe SAS «Top Finance»')],limit=1).id
            vals['credentials'] = value[5]
            vals['average_score'] = float(value[6]) if value[6] else 0
            vals['topic_specialism'] = value[7]

            contract_notes = value[8]
            if contract_notes:
                contract_notes = contract_notes.replace("\n","<br/>")
            
            vals['contract_notes'] = contract_notes


            
            if 'instructor_id' in vals:
                evaluation = self.env['mrm.instructor.evaluation'].search([('instructor_id','=',vals['instructor_id']),('name','=',vals['name'])])

                if not evaluation:               
                    evaluation = self.env['mrm.instructor.evaluation'].create(vals)
                else:
                    evaluation.write(vals)


    def import_iht_evaluations(self):
        wb_obj = load_workbook(filename = 'data/filestore/morgan13-production-1206400/d6/' + self.name)
        wsheet = wb_obj['Sheet1']
        dataDict = {}

        counter = 0
        for key, *values in wsheet.iter_rows():
            key = key.value + str(counter) if key.value else str(counter)
            dataDict[key] = [v.value for v in values]
            counter = counter + 1


        for key, value in dataDict.items():
            vals = {}
            if value[0]:
                vals['partner_id'] = self.env['res.partner'].search([('name','=',value[0])],limit=1).id
            vals['name'] = value[1]
            vals['code'] = value[2]
            date = value[3]
            try:
                date = date.replace("/", "-")
                date = datetime.strptime(date, "%d-%m-%Y")


            except:
                date = date

            vals['type'] = 'iht'
            vals['date'] = date

            if value[5]:
                employee = self.env['hr.employee'].search([('mrm_id','=',int(value[5]))])
                vals['instructor_id'] = employee.id

            if value[6]:
                company = self.env['res.company'].search([('name','=',value[6])])
                if not company:
                    company = self.env['res.company'].search([('name','ilike',value[6])])
                vals['region_id'] = company.id

            term = self.env['term'].search([('name','=',value[7]),('mrm_id','>',0)],limit=1)
            vals['term_id'] = term.id

            if value[8]:
                vals['average_score'] = float(value[8])

            
            vals['active_students'] = int(value[9])

            filled_evaluation = value[11]
            if filled_evaluation == '#DIV/0!':
                filled_evaluation = 0
            else:
                filled_evaluation = float(filled_evaluation)

            vals['filled_evaluation'] = filled_evaluation

            session = self.env['event.track'].search([('mrm_id','=', int(value[12]))])
            vals['session_id'] = session.id

            evaluation_form = value[13]
            if evaluation_form:
                evaluation_form = evaluation_form.replace("\n","<br/>")
            vals['evaluation_form'] = evaluation_form

            #evaluation = self.env['mrm.instructor.evaluation'].search([('mrm_id','=',vals['mrm_id'])])

            #if not evaluation:                
            evaluation = self.env['mrm.instructor.evaluation'].create(vals)

            #else:
                #evaluation.write(vals)

        self.is_processed = True


    def import_iht_registrations(self):
        offset = self.env['ir.config_parameter'].browse(75)
        value = offset.value
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        new_value = int(value) + 5000
        if new_value > len(reader):
            new_value = len(reader)

        for i in range(int(value), new_value):
            lines = reader[i].split('\t')
            vals = {}

            parent = False
            if lines[2]:
                parent = self.env['res.partner'].search([('mrm_id','=',int(lines[1]))])
                if not parent:
                    parent = self.env['res.partner'].create({
                        'company_type': 'company',
                        'mrm_id': lines[1],
                        'name': lines[0]
                        })


            if lines[3]:
                partner = self.env['res.partner'].search([('mrm_id','=',int(lines[3]))])
                if not partner:
                    partner = self.env['res.partner'].create({
                        'company_type': 'person',
                        'mrm_id': lines[3],
                        'name': lines[2],
                        'parent_id': parent.id,
                        })
                else:
                    partner.write({'parent_id': parent.id})
                
                vals['partner_id'] = partner.id
            
            vals['registration_number'] = lines[4]
            
            session = self.env['event.track'].search([('mrm_id','=',int(lines[6]))])
            #vals['session_id'] = session.id
            vals['event_id'] = session.event_id.id

            if lines[15]:
                vals['state'] = 'cancel'

            if 'partner_id' in vals:
                registration = self.env['event.registration'].search([('partner_id','=',vals['partner_id']),('event_id','=',vals['event_id'])])
                if not registration:
                    registration = self.env['event.registration'].create(vals)

                else:
                    registration.write(vals)

        offset.write({'value': str(new_value)})


    def import_commercial_contacts(self):
        offset = self.env['ir.config_parameter'].browse(65)
        value = offset.value
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        gender = {
        'Male': 'male',
        'Female': 'female',
        }

        new_value = int(value) + 5000
        if new_value > len(reader):
            new_value = len(reader)

        for i in range(int(value), new_value):
            lines = reader[i].split('\t')
            vals = {}
            vals['from_excel'] = True
            vals['mrm_id'] = int(lines[7])
            parent = self.env['res.partner'].search([('mrm_id','=',lines[0])])
            if not parent:
                parent = self.env['res.partner'].create({
                    'mrm_id': lines[0],
                    'from_excel': True,
                    'name': lines[1],
                    'company_type': 'company'
                    })
            else:
                parent.write({
                    'mrm_id': lines[0],
                    'from_excel': True,
                    'name': lines[1],
                    'company_type': 'company'
                    })
            vals['parent_id'] = parent.id



            vals['name'] = lines[8]
            vals['first_name'] = lines[9]
            vals['last_name'] = lines[11]
            vals['email'] = lines[12]
            vals['mobile'] = lines[13]
            vals['phone'] = lines[14]
            vals['company_type'] = 'person' if lines[6] == 'TRUE' else 'company'

            if lines[15]:
                vals['gender'] = gender.get(lines[15])

            vals['zip'] = lines[18]
            if lines[19]:
                company = self.env['res.company'].search([('name','=',lines[19])])
                if not company:
                    company = self.env['res.company'].search([('name','ilike',lines[19])],limit=1)
                
                vals['company_id'] = company.id

            vals['street'] = lines[20]
            vals['city'] = lines[21]

            if lines[23]:
                industry = self.env['res.partner.industry'].search([('name','=',lines[23])])
                if not industry:
                    industry = self.env['res.partner.industry'].create({
                        'name': lines[23]
                        })

                vals['industry_id'] = industry.id

            if lines[24]:
                country = self.env['res.country'].search([('code','=',lines[24])])
                vals['country_id'] = country.id if country else False

            vals['fax'] = lines[25]

            if lines[26]:
                education = self.env['education.level'].search([('name','=',lines[26])])
                if not education:
                    education = self.env['education.level'].create({
                        'name': lines[26]
                        })
                vals['education_level'] = education.id

            vals['university'] = lines[27]
            vals['function'] = lines[30]

            partner = self.env['res.partner'].search([('email','=',vals['email'])])
            if not partner:
                partner = self.env['res.partner'].search([('mrm_id','=',vals['mrm_id'])])
                if not partner:
                    partner = self.env['res.partner'].create(vals)

                else:
                    partner.write(vals)

            else:
                partner.write(vals)

        offset.write({'value': str(new_value)})


    def import_opportunities(self):
        offset = self.env['ir.config_parameter'].browse(76)
        value = offset.value
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        study_format = {
        'In-House Training': 'inhouse',
        'Live Class': 'live',
        'Live Class, In-House Training, Workshop': 'multiple',
        'Live Class, Live Online': 'multiple',
        'Live Class, Self-Study (packages that help candidates pass)': 'multiple',
        'Live Class, Self-Study (packages that help candidates pass), Stand-Alone (individual products)': 'multiple',
        'Live Class, Self-Study (packages that help candidates pass), Live Online': 'multiple',
        'Live Online': 'liveonline',
        'Other': 'other',
        'Self-Study (packages that help candidates pass)': 'self',
        'Self-Study (packages that help candidates pass), Live Online': 'self',
        'Stand-Alone (individual products)': 'standalone',
        'Workshop': 'workshop',
        'Workshop, Other': 'workshop',
        'Workshops': 'workshop',
        }

        stage = {
        'Submissions': 1,
        'Enquiries': 1,
        '1. Opportunity Initiation': 1,
        '2. General Inquiry': 1,
        '3. A - Course Inquiry': 2,
        '3. B - Interest unavailable': 16,
        '3. C - Marketing Event': 2,
        '4. A - Qualification': 3,
        ' 4. B - Nurturing Cycle': 3,
        '5. A - Prospecting': 3,
        '5. B - Non-Qualified': 16,
        '6. A - Current Candidate': 11,
        '6. B - Archived': 14,
        '6. C - Dead': 16,
        '7. A - Drop-Out': 16,
        '7. B - Alumni': 12,
        '7. C - Duplicate Opportunity': 16,
        }


        new_value = int(value) + 5000
        if new_value > len(reader):
            new_value = len(reader)

        for i in range(int(value), new_value):
            lines = reader[i].split('\t')
            vals = {}
            vals['mrm_id'] = int(lines[0])
            vals['name'] = lines[1]
            company = self.env['res.company'].search([('name','=',lines[4])])
            if not company:
                company = self.env['res.company'].search([('name','ilike',lines[4])])

            vals['company_id'] = company.id
            

            """user = self.env['res.users'].search([('name','=',lines[5])])
            if not user:
                user = self.env['res.users'].sudo().create({
                    'name': lines[5]
                    })
            vals['user_id'] = user.id"""

            vals['mrm_user'] = lines[5]
            
            if lines[6]:
                term = self.env['term'].search([('name','=',lines[6])])
                if not term:
                    term = self.env['term'].create({
                        'name': lines[6]
                        })

                vals['reporting_period'] = term.id

            vals['mrm_age'] = int(lines[7])

            if lines[8]:
                program = self.env['program'].search([('name','=',lines[8]),('mrm_id','>',0)],limit=1)
                if not program:
                    program = self.env['program'].create({
                        'name': lines[8]
                        })
                vals['program_id'] = program.id

            if lines[9]:
                vals['study_format'] = study_format.get(lines[9])

            factor_of_decision = self.env['factor.decision'].search([('name','=',lines[10])])
            if not factor_of_decision:
                factor_of_decision = self.env['factor.decision'].create({
                    'name': lines[10]
                    })

            vals['factor_of_decision'] = factor_of_decision.id

            if lines[11]:
                vals['probability'] = float(lines[11].replace("%",""))

            if lines[12] == 'Lost':
                vals['stage_id'] = 16

            vals['stage_id'] = stage.get(lines[13]) if 'stage_id' not in vals else 16
            vals['type'] = 'opportunity'
            vals['migration_type'] = 'opportunity'

            create_date = lines[14]
            create_date = create_date.replace("/", "-")
            create_date = create_date[:-2] + "20" + create_date[-2:]
            #raise UserError(create_date)
            create_date = datetime.strptime(create_date, "%m-%d-%Y")
            vals['mrm_create_date'] = create_date

            vals['mrm_invoice_total'] = float(lines[16])
            vals['mobile'] = lines[17]
            vals['email_from'] = lines[18]

            if lines[18]:
                partner = self.env['res.partner'].search([('email','=',lines[18])])
                vals['partner_id'] = partner.id

            description = ''
            if lines[19]:
                description += lines[19]

            if lines[20]:
                description = description + '\n' + lines[20]

            if lines[21]:
                description = description + '\n' + lines[21]

            if lines[22]:
                description = description + '\n' + lines[22]

            vals['description'] = description

            vals['partner_name'] = lines[23]

            if lines[24]:
                vals['country_id'] = self.env['res.country'].search([('name','=',lines[24])]).id

            if lines[25]:
                opp_source = self.env['utm.medium'].search([('name','=',lines[25])])
                if not opp_source:
                    opp_source = self.env['utm.medium'].create({
                        'name': lines[25]
                        })

                vals['medium_id'] = opp_source.id

            if lines[26]:
                activity = self.env['utm.campaign'].search([('name','=',lines[26])])
                if not activity:
                    activity = self.env['utm.campaign'].create({
                        'name': lines[26]
                        })

                vals['campaign_id'] = activity.id

            if lines[27]:
                source = self.env['utm.source'].search([('name','=',lines[27])])
                if not source:
                    source = self.env['utm.source'].create({
                        'name': lines[27]
                        })

                vals['source_id'] = source.id

            if lines[28]:
                industry = self.env['res.partner.industry'].search([('name','=',lines[28])])
                if not industry:
                    industry = self.env['res.partner.industry'].create({
                        'name': lines[28]
                        })

                vals['industry_id'] = industry.id

            if lines[29]:
                date = lines[29]
                date = date.replace("/", "-")
                date = date[:-2] + "20" + date[-2:]
                date = datetime.strptime(date, "%m-%d-%Y")
                vals['date_deadline'] = date

            opportunity = self.env['crm.lead'].search([('mrm_id','=',vals['mrm_id']),('migration_type','=',vals['migration_type'])],limit=1)
            if not opportunity:
                self.env['crm.lead'].create(vals)

            else:
                opportunity.write(vals)

        offset.write({'value': str(new_value)})

    def fix_opportunities_companies(self):
        offset = self.env['ir.config_parameter'].browse(76)
        value = offset.value
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        new_value = int(value) + 5000
        if new_value > len(reader):
            new_value = len(reader)

        for i in range(int(value), new_value):
            lines = reader[i].split('\t')
            vals = {}
            vals['mrm_id'] = int(lines[0])
            vals['name'] = lines[1]
            company = self.env['res.company'].search([('name','=',lines[4])])
            if not company:
                company = self.env['res.company'].search([('name','ilike',lines[4])])

            if not company and lines[4] == 'Amman':
                company = self.env['res.company'].search([('name','=','Morgan International Jordan')])

            if not company and lines[4] == 'Abu Dhabi':
                company = self.env['res.company'].search([('name','=','Morgan International Training Institute')])

            if not company and lines[4] == 'Bangalore Ample':
                company = self.env['res.company'].search([('name','=','Ample Training Institute Bangalore Pvt. Ltd.')])
 
            vals['company_id'] = company.id

            opportunity = self.env['crm.lead'].search([('mrm_id','=',vals['mrm_id']),('migration_type','=','opportunity'),('company_id','=',False)])
            if opportunity:
                try:
                    self.env.cr.execute("update crm_lead set company_id = " + str(company.id) + " where id = " + str(opportunity.id))

                except:
                    raise UserError(lines[0])

        offset.write({'value': str(new_value)})

    def import_opportunities(self):
        offset = self.env['ir.config_parameter'].browse(48)
        value = offset.value
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        study_format = {
        'In-House Training': 'inhouse',
        'Live Class': 'live',
        'Live Class, In-House Training, Workshop': 'multiple',
        'Live Class, Live Online': 'multiple',
        'Live Class, Self-Study (packages that help candidates pass)': 'multiple',
        'Live Class, Self-Study (packages that help candidates pass), Stand-Alone (individual products)': 'multiple',
        'Live Class, Self-Study (packages that help candidates pass), Live Online': 'multiple',
        'Live Online': 'liveonline',
        'Other': 'other',
        'Self-Study (packages that help candidates pass)': 'self',
        'Self-Study (packages that help candidates pass), Live Online': 'self',
        'Stand-Alone (individual products)': 'standalone',
        'Workshop': 'workshop',
        'Workshop, Other': 'workshop',
        'Workshops': 'workshop',
        'In-House Training, Live Online': 'inhouse',
        'Live Class, Workshop': 'live',
        'Live Class, Other': 'live',
        'Self-Study (packages that help candidates pass), Other': 'self',
        'Other, Live Online': 'liveonline',
        'Stand-Alone (individual products), In-House Training, Workshop': 'standalone',
        'Live Class, Workshop, Other': 'live',
        'Live Class, Self-Study (packages that help candidates pass), Other': 'live',
        'Live Class, In-House Training': 'live',
        'In-House Training, Workshop': 'inhouse'
        }

        stage = {
        'Submissions': 1,
        'Enquiries': 1,
        '1. Opportunity Initiation': 1,
        '2. General Inquiry': 1,
        '3. A - Course Inquiry': 2,
        '3. B - Interest unavailable': 10,
        '3. C - Marketing Event': 2,
        '4. A - Qualification': 3,
        '4. B - Nurturing Cycle': 3,
        '5. A - Prospecting': 3,
        '5. B - Non-Qualified': 10,
        '6. A - Current Candidate': 7,
        '6. B - Archived': 9,
        '6. C - Dead': 10,
        '7. A - Drop-Out': 10,
        '7. B - Alumni': 8,
        '7. C - Duplicate Opportunity': 10,
        'Active Student': 7,
        'Archived': 9,
        'Decide': 6,
        'Evaluate Option': 3,
        'New': 1,
        'Lost': 10,
        'Seek Advice': 2
        }

        gender = {
        'Male': 'male',
        'Female': 'female',
        }

        new_value = int(value) + 1000
        if new_value > len(reader):
            new_value = len(reader)

        for i in range(int(value), new_value):

            lines = reader[i].split('\t')
            vals = {}



            vals['mrm_id'] = int(lines[0])
            vals['name'] = lines[1]

            company_type = 'person'
            account_type = 'b2c'
            partner = False

            if lines[4] == 'FALSE':
                company_type = 'company'
                account_type = 'b2b'    
            
            if account_type == 'b2b':
                if lines[6]:
                    parent_company = self.env['res.partner'].search([('mrm_id','=',int(lines[6])),('is_mrm_contact','=',False)])
                    if len(parent_company) > 1:
                        parent_company = self.env['res.partner'].search([('mrm_id','=',int(lines[6])),('name','=',lines[5]),('is_mrm_contact','=',False)])
                    
                    if not parent_company:
                        parent_company = self.env['res.partner'].create({
                            'mrm_id': int(lines[6]),
                            'name': lines[5],
                            'account_type': account_type,
                            'company_type': company_type,
                            })
                    else:
                        parent_company.write({
                            'mrm_id': int(lines[6]),
                            'name': lines[5],
                            'account_type': account_type,
                            'company_type': company_type,
                            })

                    if lines[8]:
                        partner = self.env['res.partner'].search([('mrm_id','=',int(lines[8])),('is_mrm_contact','=',True)],limit=1)
                        if not partner:
                            partner = self.env['res.partner'].create({
                                'mrm_id': int(lines[8]),
                                'name': lines[7].strip(),
                                'account_type': 'b2c',
                                'company_type': 'person',
                                'parent_id': parent_company.id,
                                'gender': gender.get(lines[9]),
                                'is_mrm_contact': True,
                                })
                        else:
                            partner.write({
                                'mrm_id': int(lines[8]),
                                'name': lines[7].strip(),
                                'account_type': 'b2c',
                                'company_type': 'person',
                                'parent_id': parent_company.id,
                                'gender': gender.get(lines[9]),
                                'is_mrm_contact': True
                                })
                        partner = partner.id

            elif account_type == 'b2c':
                if lines[8]:
                    partner = self.env['res.partner'].search([('mrm_id','=',int(lines[8])),('is_mrm_contact','=',True)],limit=1)
                    if not partner:
                        partner = self.env['res.partner'].create({
                            'mrm_id': int(lines[8]),
                            'name': lines[7].strip(),
                            'account_type': 'b2c',
                            'company_type': 'person',
                            'gender': gender.get(lines[9]),
                            'is_mrm_contact': True
                            })

                    else:
                        partner.write({
                            'mrm_id': int(lines[8]),
                            'name': lines[7].strip(),
                            'account_type': 'b2c',
                            'company_type': 'person',
                            'gender': gender.get(lines[9]),
                            'is_mrm_contact': True
                            })
                    partner = partner.id

            vals['partner_id'] = partner

            if lines[10]:
                company = self.env['res.company'].search([('name','=',lines[10].strip())])
                if not company:
                    company = self.env['res.company'].search([('name','ilike',lines[10].strip())])

                vals['company_id'] = company.id

            vals['mrm_user'] = lines[11]


            if lines[11]:
                user = self.env['res.users'].search([('name','=',lines[11]),('active','in',[True,False])],limit=1)
                if not user:
                    user = self.env['res.users'].search([('name','=',lines[11]),('active','in',[True,False])],limit=1)

                vals['user_id'] = user.id
            
            if lines[12]:
                term = self.env['term'].search([('name','=',lines[12].strip())],limit=1)
                if not term:
                    term = self.env['term'].create({
                        'name': lines[12].strip()
                        })

                vals['reporting_period'] = term.id

            vals['mrm_age'] = int(lines[13])

            if lines[14]:
                program = self.env['program'].search([('name','=',lines[14].strip()),('mrm_id','>',0)],limit=1)
                if not program:
                    program = self.env['program'].search([('name','=',lines[14].strip())],limit=1)

                if not program:
                    program = self.env['program'].create({
                        'name': lines[14].strip()
                        })
                
                vals['program_id'] = program.id

            if lines[15]:
                vals['study_format'] = study_format.get(lines[15])

            factor_of_decision = self.env['factor.decision'].search([('name','=',lines[16].strip())])
            if not factor_of_decision:
                factor_of_decision = self.env['factor.decision'].create({
                    'name': lines[16].strip()
                    })

            vals['factor_of_decision'] = factor_of_decision.id

            if lines[17]:
                vals['probability'] = float(lines[17].replace("%",""))

            if lines[18] == 'Lost':
                vals['stage_id'] = 10

            if lines[19] != 'Website Order' and lines[19] != 'Website Order - Offline Payment':
                vals['stage_id'] = stage[lines[19].strip()] if 'stage_id' not in vals else 10
            
            vals['type'] = 'opportunity'
            vals['migration_type'] = 'opportunity'

            create_date = lines[20]
            #create_date = create_date.replace("/", "-")
            #create_date = create_date[:-2] + "20" + create_date[-2:]
            #raise UserError(create_date)
            create_date = datetime.strptime(create_date, "%Y-%m-%d")
            vals['mrm_create_date'] = create_date

            vals['mrm_invoice_total'] = float(lines[22])
            vals['mobile'] = lines[23]
            vals['email_from'] = lines[24]

            description = ''
            if lines[25]:
                description += lines[25]

            if lines[26]:
                description = description + '\n' + lines[26]

            if lines[27]:
                description = description + '\n' + lines[27]

            if lines[28]:
                description = description + '\n' + lines[28]

            vals['description'] = description

            vals['partner_name'] = lines[29]

            if lines[30]:
                vals['country_id'] = self.env['res.country'].search([('name','=',lines[30].strip())]).id

            if lines[31]:
                opp_source = self.env['utm.medium'].search([('name','=',lines[31].strip())])
                if not opp_source:
                    opp_source = self.env['utm.medium'].create({
                        'name': lines[31].strip()
                        })

                vals['medium_id'] = opp_source.id

            if lines[32]:
                activity = self.env['utm.campaign'].search([('name','=',lines[32].strip())])
                if not activity:
                    activity = self.env['utm.campaign'].create({
                        'name': lines[32].strip()
                        })

                vals['campaign_id'] = activity.id

            if lines[33]:
                source = self.env['utm.source'].search([('name','=',lines[33].strip())])
                if not source:
                    source = self.env['utm.source'].create({
                        'name': lines[33].strip()
                        })

                vals['source_id'] = source.id

            if lines[34]:
                industry = self.env['res.partner.industry'].search([('name','=',lines[34].strip())])
                if not industry:
                    industry = self.env['res.partner.industry'].create({
                        'name': lines[34].strip()
                        })

                vals['industry_id'] = industry.id

            if lines[35]:
                date = lines[35]
                date = date.replace("/", "-")
                date = date[:-2] + "20" + date[-2:]
                date = datetime.strptime(date, "%m-%d-%Y")
                vals['date_deadline'] = date

            opportunity = self.env['crm.lead'].search([('mrm_id','=',vals['mrm_id']),('migration_type','=',vals['migration_type'])],limit=1)
            if not opportunity:
                self.env['crm.lead'].create(vals)

            else:
                opportunity.write(vals)

        offset.write({'value': str(new_value)})


    def import_enquiries(self):
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        study_format = {
        'In-House Training': 'inhouse',
        'Live Class': 'live',
        'Live Class, In-House Training, Workshop': 'multiple',
        'Live Class, Live Online': 'multiple',
        'Live Class, Self-Study (packages that help candidates pass)': 'multiple',
        'Live Class, Self-Study (packages that help candidates pass), Stand-Alone (individual products)': 'multiple',
        'Live Class, Self-Study (packages that help candidates pass), Live Online': 'multiple',
        'Live Online': 'liveonline',
        'Other': 'other',
        'Self-Study (packages that help candidates pass)': 'self',
        'Self-Study (packages that help candidates pass), Live Online': 'self',
        'Stand-Alone (individual products)': 'standalone',
        'Workshop': 'workshop',
        'Workshop, Other': 'workshop',
        'Workshops': 'workshop',
        }

        stage = {
        'Submissions': 1,
        'Enquiries': 1,
        '1. Opportunity Initiation': 1,
        '2. General Inquiry': 1,
        '3. A - Course Inquiry': 2,
        '3. B - Interest unavailable': 16,
        '3. C - Marketing Event': 2,
        '4. A - Qualification': 3,
        ' 4. B - Nurturing Cycle': 3,
        '5. A - Prospecting': 3,
        '5. B - Non-Qualified': 16,
        '6. A - Current Candidate': 11,
        '6. B - Archived': 14,
        '6. C - Dead': 16,
        '7. A - Drop-Out': 16,
        '7. B - Alumni': 12,
        '7. C - Duplicate Opportunity': 16,
        }

        for row in reader:
            lines = row.split('\t')
            vals = {}

            if lines[4] == 'FALSE':
                vals['mrm_id'] = int(lines[0])
                vals['type'] = 'opportunity'
                vals['migration_type'] = 'enquiry'

                create_date = str(lines[1])
                month = create_date.split('/')[0]

                if int(month) < 10:
                    create_date = create_date[:7].rstrip()

                else:
                    create_date = create_date[:8].rstrip()

                create_date = create_date.replace("/", "-")
                create_date = create_date[:-2] + "20" + create_date[-2:]
                #raise UserError(create_date)
                create_date = datetime.strptime(create_date, "%m-%d-%Y")
                vals['mrm_create_date'] = create_date

                partner = self.env['res.partner'].search([('mrm_id','=',lines[2])])
                if partner:
                    vals['partner_id'] = partner.id

                vals['opportunity_mrm_id'] = int(lines[5]) if lines[5] else False

                vals['mrm_user'] = lines[7]

                if lines[7]:
                    user = self.env['res.users'].search([('name','=',lines[7]),('active','in',[True,False])],limit=1)
                    if not user:
                        user = self.env['res.users'].search([('name','=',lines[7]),('active','in',[True,False])],limit=1)

                    vals['user_id'] = user.id

                company = self.env['res.company'].search([('name','=',lines[8])])

                vals['company_id'] = company.id if company else False

                vals['email_from'] = lines[9]

                vals['mrm_subject'] = lines[10]

                vals['mrm_age'] = lines[11]

                vals['mrm_body'] = lines[12]

                vals['mrm_dead_reason'] = lines[13]

                vals['name'] = lines[14]

                if lines[14]:
                    source = self.env['utm.source'].search([('name','=',lines[14])])
                    if not source:
                        source = self.env['utm.source'].create({
                            'name': lines[14]
                            })

                    vals['source_id'] = source.id

                if lines[5]:
                    vals['stage_id'] = self.env['crm.lead'].search([('mrm_id','=',int(lines[5]))]).stage_id.id

                opportunity = self.env['crm.lead'].search([('mrm_id','=',vals['mrm_id']),('migration_type','=',vals['migration_type'])])
                if not opportunity:
                    self.env['crm.lead'].create(vals)

                else:
                    opportunity.write(vals)

                if vals['mrm_body']:
                    self.env['mail.message'].search([('res_id','=',opportunity.id),('model','=','crm.lead')]).unlink()
                    self.env['mail.message'].create({
                        'res_id': opportunity.id,
                        'subject': vals['mrm_subject'],
                        'res_id': opportunity.id,
                        'email_from': vals['email_from'],
                        'model': 'crm.lead',
                        'body': vals['mrm_body'], 
                        })

    def import_enquiries_messages(self):
        wb_obj = load_workbook(filename = 'data/filestore/morgan13-stage-1493458/d6/Commercial_Migration_Odoo_Final.xlsx')
        wsheet = wb_obj['Enquiries']
        dataDict = {}

        counter = 0
        for key, *values in wsheet.iter_rows():
            key = str(counter)
            dataDict[key] = [v.value for v in values]
            counter = counter + 1


        for key, value in dataDict.items():
            vals = {}
            vals['mrm_id'] = int(value[0])
            vals['migration_type'] = 'enquiry'

            vals['email_from'] = value[9]

            vals['mrm_subject'] = value[10]

            vals['mrm_body'] = value[12]

            opportunity = self.env['crm.lead'].search([('mrm_id','=',vals['mrm_id']),('migration_type','=',vals['migration_type'])])
            #raise UserError(str(vals))
            if opportunity:
                if vals['mrm_body']:
                    self.env['mail.message'].search([('res_id','=',opportunity.id),('model','=','crm.lead')]).unlink()
                    self.env['mail.message'].create({
                        'res_id': opportunity.id,
                        'subject': vals['mrm_subject'],
                        'res_id': opportunity.id,
                        'email_from': vals['email_from'],
                        'model': 'crm.lead',
                        'body': vals['mrm_body'], 
                        })

    def import_submissions(self):
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        study_format = {
        'In-House Training': 'inhouse',
        'Live Class': 'live',
        'Live Class, In-House Training, Workshop': 'multiple',
        'Live Class, Live Online': 'multiple',
        'Live Class, Self-Study (packages that help candidates pass)': 'multiple',
        'Live Class, Self-Study (packages that help candidates pass), Stand-Alone (individual products)': 'multiple',
        'Live Class, Self-Study (packages that help candidates pass), Live Online': 'multiple',
        'Live Online': 'liveonline',
        'Other': 'other',
        'Self-Study (packages that help candidates pass)': 'self',
        'Self-Study (packages that help candidates pass), Live Online': 'self',
        'Stand-Alone (individual products)': 'standalone',
        'Workshop': 'workshop',
        'Workshop, Other': 'workshop',
        'Workshops': 'workshop',
        }

        stage = {
        'Submissions': 1,
        'Enquiries': 1,
        '1. Opportunity Initiation': 1,
        '2. General Inquiry': 1,
        '3. A - Course Inquiry': 2,
        '3. B - Interest unavailable': 16,
        '3. C - Marketing Event': 2,
        '4. A - Qualification': 3,
        ' 4. B - Nurturing Cycle': 3,
        '5. A - Prospecting': 3,
        '5. B - Non-Qualified': 16,
        '6. A - Current Candidate': 11,
        '6. B - Archived': 14,
        '6. C - Dead': 16,
        '7. A - Drop-Out': 16,
        '7. B - Alumni': 12,
        '7. C - Duplicate Opportunity': 16,
        }

        for row in reader:
            lines = row.split('\t')
            vals = {}

            if lines[4] == 'FALSE':
                campaign = False
                vals['mrm_id'] = int(lines[0])
                vals['type'] = 'opportunity'
                vals['migration_type'] = 'submission'

                vals['mrm_create_date'] = datetime.strptime(lines[1], "%Y-%m-%d")

                partner = self.env['res.partner'].search([('mrm_id','=',lines[2])])

                vals['partner_id'] = partner.id if partner else False

                vals['mrm_user'] = lines[5]
                vals['mrm_age'] = int(lines[6])
                vals['email_from'] = lines[9]

                country = self.env['res.country'].search([('name','=',lines[10])])
                vals['country_id'] = country.id if country else False

                if lines[11]:
                    company = self.env['res.company'].search([('name','=',lines[11])])
                    if company:
                        vals['company_id'] = company.id if company else False

                if lines[12]:
                    program = self.env['program'].search([('name','=',lines[12]),('mrm_id','>',0)])
                    vals['program_id'] = program.id

                vals['mrm_url'] = lines[13]

                description = ''
                if lines[14]:
                    description += lines[14]

                if lines[15]:
                    source = self.env['utm.source'].search([('name','=',lines[15])])
                    if not source:
                        source = self.env['utm.source'].create({
                            'name': lines[15]
                            })

                    vals['source_id'] = source.id

                if lines[16]:
                    campaign = self.env['utm.campaign'].search([('name','=',lines[16])])
                    if not campaign:
                        campaign = self.env['utm.campaign'].create({
                            'name': lines[16]
                            })

                    vals['campaign_id'] = campaign.id

                if lines[18]:
                    description = description + '\n' + lines[18]

                if lines[19]:
                    description = description + '\n' + lines[19]

                vals['description'] = description

                vals['name'] = 'Submission - ' + str(lines[0])

                campaign_start_date = lines[26]
                campaign_start_date = campaign_start_date.replace("/", "-")
                campaign_start_date = campaign_start_date[:-2] + "20" + campaign_start_date[-2:]
                campaign_start_date = datetime.strptime(campaign_start_date, "%m-%d-%Y")

                campaign_end_date = lines[27]
                campaign_end_date = campaign_end_date.replace("/", "-")
                campaign_end_date = campaign_end_date[:-2] + "20" + campaign_end_date[-2:]
                campaign_end_date = datetime.strptime(campaign_end_date, "%m-%d-%Y")

                campaign.write({
                    'mrm_purpose': lines[23],
                    'mrm_term_id': self.env['term'].search([('name','=',lines[24])]).id,
                    'mrm_cost': float(lines[25]),
                    'mrm_start_date': campaign_start_date,
                    'mrm_end_date': campaign_end_date
                    })




                opportunity = self.env['crm.lead'].search([('mrm_id','=',vals['mrm_id']),('migration_type','=',vals['migration_type'])])
                if not opportunity:
                    self.env['crm.lead'].create(vals)

                else:
                    opportunity.write(vals)


    def import_mrm_quotations(self):
        wb_obj = load_workbook(filename = 'data/filestore/morgan13-stage-1493458/98/Morgan International Lebanon LBP 1.xlsx')
        wsheet = wb_obj['Sheet1']
        dataDict = {}

        counter = 0
        for key, *values in wsheet.iter_rows():
            key = str(counter)
            dataDict[key] = [v.value for v in values]
            counter = counter + 1


        for key, value in dataDict.items():
            vals = {}

            vals['mrm'] = True

            if value[0]:
                vals['name'] = 'SO ' + value[0]

            if value[1]:
                date = datetime.strptime(value[1], '%b %d, %Y')
                vals['date_order'] = date

            if value[3]:
                partner = self.env['res.partner'].search([('mrm_id','=',int(value[3]))],limit=1)

                if not partner:
                    partner = self.env['res.partner'].search([('name','=',value[2])],limit=1)

                    if not partner:
                        partner = self.env['res.partner'].create({
                            'name': value[2],
                            'mrm_id': value[3],
                            'company_type': 'person',
                            'from_excel': True,
                            })

                    else:
                        partner.write({
                            'name': value[2],
                            'mrm_id': value[3],
                            'company_type': 'person',
                            'from_excel': True,
                            })

                else:
                    partner.write({
                        'name': value[2],
                        'mrm_id': value[3],
                        'company_type': 'person',
                        'from_excel': True,
                        })



                vals['partner_id'] = partner.id

            vals['other_inv_ref'] = value[4]
            vals['note'] = value[23]

            if value[20]:
                company = self.env['res.company'].search([('name','=',value[20])],limit=1)
                vals['company_id'] = company.id

                currency = self.env['res.currency'].search([('name','=',value[10])])
                pricelist = self.env['product.pricelist'].search([('company_id','=',company.id),('currency_id','=',currency.id),('name','=',company.name)])
                if not pricelist:
                    pricelist = self.env['product.pricelist'].create({
                        'company_id': company.id,
                        'name': company.name,
                        'currency_id': currency.id,
                        })
                
                vals['pricelist_id'] = pricelist.id


            if 'date_order' in vals:
                order = self.env['sale.order'].search([('name','=',vals['name'])])
                if not order:
                    order = self.env['sale.order'].create(vals)
                    order.write({'name': vals['name']})

    def import_mrm_quotations(self):

        count = self.processed_rows_count
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        new_value = count + 5000
        if new_value > len(reader):
            new_value = len(reader)

            self.is_processed = True

        for i in range(count, new_value):
            lines = reader[i].split('\t')
            vals = {}

            vals['mrm'] = True

            if lines[0]:
                vals['name'] = 'SO ' + lines[0]

            if lines[1]:
                date = datetime.strptime(lines[1], '%b %d, %Y')
                vals['date_order'] = date

            if lines[3]:
                try:
                    partner = self.env['res.partner'].search([('mrm_id','=',int(lines[3]))],limit=1)

                except:
                    raise UserError(lines)

                if not partner:
                    partner = self.env['res.partner'].search([('name','=',lines[2])],limit=1)

                    if not partner:
                        partner = self.env['res.partner'].create({
                            'name': lines[2],
                            'mrm_id': lines[3],
                            'company_type': 'person',
                            'from_excel': True,
                            })

                    else:
                        partner.write({
                            'name': lines[2],
                            'mrm_id': lines[3],
                            'company_type': 'person',
                            'from_excel': True,
                            })

                else:
                    partner.write({
                        'name': lines[2],
                        'mrm_id': lines[3],
                        'company_type': 'person',
                        'from_excel': True,
                        })



                vals['partner_id'] = partner.id

            vals['other_inv_ref'] = lines[4]
            vals['note'] = lines[23]

            if lines[20]:
                company_name = lines[20].strip()
                company = self.env['res.company'].search([('name','=',company_name)],limit=1)
                if not company:
                    company = self.env['res.company'].search([('name','ilike',company_name)],limit=1)
                vals['company_id'] = company.id

                currency = self.env['res.currency'].search([('name','=',lines[10])])
                pricelist = self.env['product.pricelist'].search([('company_id','=',company.id),('currency_id','=',currency.id),('name','=',company.name)])
                if not pricelist:
                    pricelist = self.env['product.pricelist'].search([('currency_id','=',currency.id),('name','=',company.name)])
                    
                if not pricelist:
                    try:
                        pricelist = self.env['product.pricelist'].create({
                            'company_id': company.id,
                            'name': company.name,
                            'currency_id': currency.id,
                            })
                    except:
                        raise UserError(self.name + ' - ' + str(i) + ' - ' + lines[20])
                
                vals['pricelist_id'] = pricelist.id


            if 'date_order' in vals:
                order = self.env['sale.order'].search([('name','=',vals['name'])])
                if not order:
                    if 'partner_id' in vals:
                        order = self.env['sale.order'].create(vals)
                        order.write({'name': vals['name']})

        self.write({'processed_rows_count': new_value})
        

    def import_mrm_quotation_lines(self):
        wb_obj = load_workbook(filename = 'data/filestore/morgan13-stage-1493458/98/Odoo Migration Sales Report - Beirut F19 (1).xlsx')
        wsheet = wb_obj['Sheet1']
        dataDict = {}

        counter = 0
        for key, *values in wsheet.iter_rows():
            key = str(counter)
            dataDict[key] = [v.value for v in values]
            counter = counter + 1


        for key, value in dataDict.items():
            vals = {}
            date = False
            if value[0]:
                order_name = 'SO ' + str(value[0])
                order = self.env['sale.order'].search([('name','=',order_name)])

            if value[1]:
                date = datetime.strptime(value[1], '%b %d, %Y')

                if order:
                    if value[10]:
                        company = self.env['res.company'].search([('name','=',value[20])],limit=1)

                        if value[9]:
                            product = self.env['product.template'].search([('mrm_id','=',int(value[9]))],limit=1)
                            if not product:
                                product = self.env['product.template'].create({
                                    'mrm_id': int(value[9]),
                                    'name': value[8],
                                    'type': 'product' if value[22] == 'TRUE' else 'service'
                                    })
                            else:
                                product.write({
                                    'name': value[8],
                                    'type': 'product' if value[22] == 'TRUE' else 'service'
                                    })
                            
                            product_variant = product.product_variant_id
                            vals['order_id'] = order.id
                            vals['product_id'] = product_variant.id
                            vals['product_uom'] = product_variant.uom_id.id
                            vals['name'] = value[8]
                            vals['product_uom_qty'] = float(value[12]) if value[12] else 0
                            vals['discount'] = float(value[13]) if value[13] else 0
                            vals['price_unit'] = float(value[11]) if value[11] else 0

                            tax = False
                            if value[16] and float(value[16]) > 0:
                                tax = self.env['account.tax'].search([('mrm_id','>',0),('company_id','=',company.id),('amount','=',float(value[16]))])
                                if not tax:
                                    tax = self.env['account.tax'].search([('mrm_id','>',0),('company_id','=',company.parent_id.id),('amount','=',float(value[16]))])

                                vals['tax_id'] = [(6, 0, [tax.id])] if tax else [(5, 0, 0)]

                            term = self.env['term'].search([('name','=',value[19])])
                            vals['term_id'] = term.id

                            if value[21]:
                                event = self.env['event.event'].search([('mrm_id','=',int(value[21]))])
                                vals['event_id'] = event.id
                            
                            order_line = self.env['sale.order.line'].create(vals)

                    order.action_confirm()
                    order.write({'date_order': date})


    def import_mrm_quotation_lines(self):
        count = self.quotation_lines_row_count
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        new_value = count + 5000
        if new_value > len(reader):
            new_value = len(reader)
            

            self.lines_processed = True

        for i in range(count, new_value):
            lines = reader[i].split('\t')
            vals = {}
            date = False
            if lines[0]:
                order_name = 'SO ' + str(lines[0])
                order = self.env['sale.order'].search([('name','=',order_name)])

            if lines[1]:
                date = datetime.strptime(lines[1], '%b %d, %Y')

                if order:
                    if lines[10]:
                        company = self.env['res.company'].search([('name','=',lines[20])],limit=1)

                        if lines[9]:
                            product = self.env['product.template'].search([('mrm_id','=',int(lines[9]))],limit=1)
                            if not product:
                                product = self.env['product.template'].create({
                                    'mrm_id': int(lines[9]),
                                    'name': lines[8],
                                    'type': 'product' if lines[22] == 'TRUE' else 'service'
                                    })
                            else:
                                product.write({
                                    'name': lines[8],
                                    'type': 'product' if lines[22] == 'TRUE' else 'service'
                                    })
                            
                            product_variant = product.product_variant_id
                            vals['order_id'] = order.id
                            vals['product_id'] = product_variant.id
                            vals['product_uom'] = product_variant.uom_id.id
                            vals['name'] = lines[8]
                            vals['product_uom_qty'] = float(lines[12]) if lines[12] else 0
                            vals['discount'] = float(lines[13]) if lines[13] else 0
                            vals['price_unit'] = float(lines[11]) if lines[11] else 0

                            tax = False
                            if lines[16] and float(lines[16]) > 0:
                                tax = self.env['account.tax'].search([('mrm_id','>',0),('company_id','=',company.id),('amount','=',float(lines[16]))])
                                if not tax:
                                    tax = self.env['account.tax'].search([('mrm_id','>',0),('company_id','=',company.parent_id.id),('amount','=',float(lines[16]))])

                                vals['tax_id'] = [(6, 0, [tax.id])] if tax else [(5, 0, 0)]

                            term = self.env['term'].search([('name','=',lines[19])])
                            vals['term_id'] = term.id
                            if not term:
                                term = self.env['term'].create({
                                    'name': lines[19]
                                    })

                            if lines[21]:
                                event = self.env['event.event'].search([('mrm_id','=',int(lines[21]))])
                                vals['event_id'] = event.id
                            
                            order_line = self.env['sale.order.line'].create(vals)

                    order.action_confirm()
                    order.write({'date_order': date})

        self.write({'quotation_lines_row_count': new_value})

    def fix_invoice_and_quotation_dates(self):
        wb_obj = load_workbook(filename = 'data/filestore/morgan13-stage-1493458/98/Odoo Migration Sales Report - Beirut F19 (1).xlsx')
        wsheet = wb_obj['Sheet1']
        dataDict = {}

        counter = 0
        for key, *values in wsheet.iter_rows():
            key = str(counter)
            dataDict[key] = [v.value for v in values]
            counter = counter + 1
            


        for key, value in dataDict.items():
            vals = {}
            if value[0]:
                invoice = self.env['account.move'].search([('name','=',value[0])])
                order_name = 'SO ' + str(value[0])
                order = self.env['sale.order'].search([('name','=',order_name)])
            
            if value[1]:
                date = datetime.strptime(value[1], '%b %d, %Y')
                invoice.write({'invoice_date': date})
                order.write({'date_order': date})


    def import_receipts(self):
        wb_obj = load_workbook(filename = 'data/filestore/morgan13-stage-1493458/98/Receipt Report - Beirut F19.xlsx')
        wsheet = wb_obj['Sheet1']
        dataDict = {}

        payment_types = {
        'Bank Transfer': 'transfer',
        'Cash': 'cash',
        'Cheque': 'cheque',
        'Converted FP (Void)': 'cash',
        'Credit/Debit Card': 'credit_card',
        }

        counter = 0
        for key, *values in wsheet.iter_rows():
            key = str(counter)
            dataDict[key] = [v.value for v in values]
            counter = counter + 1


        for key, value in dataDict.items():
            vals = {}
            vals['from_mrm'] = True
            vals['partner_type'] = 'customer'
            vals['payment_method_id'] = self.env.ref('account.account_payment_method_manual_in').id

            partner = self.env['res.partner'].search([('mrm_id','=',value[1])])
            vals['partner_id'] = partner.id
            vals['name'] = value[2]

            payment_date = value[3]
            payment_date = payment_date.replace("/", "-")
            payment_date = payment_date[:-2] + "20" + payment_date[-2:]
            payment_date = datetime.strptime(payment_date, "%m-%d-%Y")
            vals['payment_date'] = payment_date

            
            if value[4]:
                date_cleared = value[4]
                date_cleared = date_cleared.replace("/", "-")
                date_cleared = date_cleared[:-2] + "20" + date_cleared[-2:]
                date_cleared = datetime.strptime(date_cleared, "%m-%d-%Y")
                vals['date_cleared'] = date_cleared
            
            #vals['invoice_ids'] = [(6, 0, [invoice.id])]

            invoice = self.env['account.move'].search([('name','=',value[5])])
            vals['company_id'] = invoice.company_id.id


            
            vals['payment_mode'] = payment_types[value[6]]

            if vals['payment_mode'] == 'cash':
                journal = self.env['account.journal'].search([('company_id','=',company_id),('type','=','cash')],limit=1)

            else:
                journal = self.env['account.journal'].search([('company_id','=',company_id),('type','=','bank')],limit=1)

            vals['journal_id'] = journal.id

            vals['payment_ref'] = value[7]

            currency = self.env['res.currency'].search([('name','=',value[8])])
            vals['currency_id'] = currency.id

            amount = float(value[9]) if value[8] != 'USD' else float(value[10])
            vals['amount'] = -amount if amount < 0 else amount

            vals['payment_type'] = 'inbound' if amount >= 0 else 'outbound'

            vals['comment'] = value[11]

            payment = self.env['account.payment'].search([('name','=',vals['name']),('from_mrm','=',True)])
            if not payment:
                payment = self.env['account.payment'].create(vals)
                payment.post()

            elif payment.state == 'posted':
                payment.action_draft()
                amount = payment.amount
                vals['amount'] += amount
                payment.write(vals)
                payment.post()


            register_payment = self.env['register.payment'].search([('payment_id','=',payment.id)])
            if not register_payment:
                register_payment = self.env['register.payment'].create({
                    'payment_id': payment.id,
                    'partner_id': vals['partner_id'],
                    'from_mrm': True,
                    })


            payment_line = self.env['register.payment.line'].search([('move_id','=',invoice.id),('payment_registration_id','=',register_payment.id)])
            if not payment_line:
                allocated_amount = float(value[9])
                allocated_amount = -allocated_amount if allocated_amount < 0 else allocated_amount
                payment_line = self.env['register.payment.line'].create({
                    'payment_registration_id': register_payment.id,
                    'move_id': invoice.id,
                    'amount': allocated_amount,
                    'allocated_amount': allocated_amount,
                    'partner_id': vals['partner_id'],
                    })

    def import_receipts(self):
        payment_types = {
        'Bank Transfer': 'transfer',
        'Cash': 'cash',
        'Cheque': 'cheque',
        'Converted FP (Void)': 'cash',
        'Credit/Debit Card': 'credit_card',
        }
        count = self.processed_rows_count
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        new_value = count + 1000
        if new_value > len(reader):
            new_value = len(reader)

            


            self.is_processed = True
            

        for i in range(count, new_value):
            lines = reader[i].split('\t')
            vals = {}
            vals['from_mrm'] = True
            vals['partner_type'] = 'customer'
            vals['payment_method_id'] = self.env.ref('account.account_payment_method_manual_in').id

            if lines[1]:
                partner = self.env['res.partner'].search([('mrm_id','=',lines[1]),('is_mrm_contact','=',False)],limit=1)
                if not partner:
                    partner = self.env['res.partner'].create({
                        'mrm_id': lines[1],
                        'name': lines[0]
                        })


                vals['partner_id'] = partner.id
                vals['name'] = lines[2]


                payment_date = lines[3]
                payment_date = payment_date.replace("/", "-")
                payment_date = payment_date[:-2] + "20" + payment_date[-2:]
                payment_date = datetime.strptime(payment_date, "%m-%d-%Y")
                vals['payment_date'] = payment_date

                
                if lines[4]:
                    date_cleared = lines[4]
                    date_cleared = date_cleared.replace("/", "-")
                    date_cleared = date_cleared[:-2] + "20" + date_cleared[-2:]
                    date_cleared = datetime.strptime(date_cleared, "%m-%d-%Y")
                    vals['date_cleared'] = date_cleared
                
                #vals['invoice_ids'] = [(6, 0, [invoice.id])]

                invoice = self.env['account.move'].search([('name','=',lines[5])])

                company = self.env['res.company'].search([('name','=',lines[13].strip())])
                if not company:
                    company = self.env['res.company'].search([('name','ilike',lines[13].strip())])
                vals['company_id'] = company.id


                
                vals['payment_mode'] = payment_types.get(lines[6])

                if vals['payment_mode'] == 'cash':
                    journal = self.env['account.journal'].search([('company_id','=',company.id),('type','=','cash')],limit=1)

                else:
                    journal = self.env['account.journal'].search([('company_id','=',company.id),('type','=','bank')],limit=1)

                vals['journal_id'] = journal.id

                vals['payment_ref'] = lines[7]

                currency = self.env['res.currency'].search([('name','=',lines[8])])
                vals['currency_id'] = currency.id

                amount = float(lines[9]) if lines[8] != 'USD' else float(lines[10])
                vals['amount'] = -amount if amount < 0 else amount

                vals['payment_type'] = 'inbound' if amount >= 0 else 'outbound'

                vals['comment'] = lines[11]

                payment = self.env['account.payment'].search([('name','=',vals['name']),('from_mrm','=',True),('partner_id','=',False),('state','!=','cancelled')])
                if payment:
                    payment.action_draft()
                    payment.write(vals)
                    payment.post()
                
                else:
                    payment = self.env['account.payment'].search([('name','=',vals['name']),('from_mrm','=',True),('partner_id','=',vals['partner_id']),('state','!=','cancelled')])
                    if not payment:
                        payment = self.env['account.payment'].create(vals)
                        try:
                            payment.post()

                        except:
                            raise UserError(str(vals))

                """elif payment.state == 'posted':
                    payment.action_draft()
                    amount = payment.amount
                    vals['amount'] += amount
                    payment.write(vals)
                    try:
                        payment.post()

                    except:
                        raise UserError(lines[2])"""


                register_payment = self.env['register.payment'].search([('payment_id','=',payment.id)])
                if not register_payment:
                    register_payment = self.env['register.payment'].create({
                        'payment_id': payment.id,
                        'partner_id': vals['partner_id'],
                        'from_mrm': True,
                        })



                payment_line = self.env['register.payment.line'].search([('move_id','=',invoice.id),('payment_registration_id','=',register_payment.id)])
                if not payment_line:
                    allocated_amount = float(lines[9])
                    allocated_amount = -allocated_amount if allocated_amount < 0 else allocated_amount
                    payment_line = self.env['register.payment.line'].create({
                        'payment_registration_id': register_payment.id,
                        'move_id': invoice.id,
                        'amount': allocated_amount,
                        'allocated_amount': allocated_amount,
                        'partner_id': vals['partner_id'],
                        })

                else:
                    allocated_amount = float(lines[9])
                    allocated_amount = -allocated_amount if allocated_amount < 0 else allocated_amount
                    payment_line.write({
                        'payment_registration_id': register_payment.id,
                        'move_id': invoice.id,
                        'amount': allocated_amount,
                        'allocated_amount': allocated_amount,
                        'partner_id': vals['partner_id'],
                        })

            self.write({'processed_rows_count': new_value})

    def fix_event_dates(self):
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()
        for row in reader:
            lines = row.split('\t')
            vals = {}
            vals['mrm_id'] = lines[1]
            date_begin = lines[8].replace("/","-")
            date_begin = date_begin + " 00:00:00"

            date_end = lines[9].replace("/","-")
            date_end = date_end + " 00:00:00"
            vals['date_begin'] = datetime.strptime (date_begin, "%d-%m-%Y %H:%M:%S")
            vals['date_end'] = datetime.strptime (date_end, "%d-%m-%Y %H:%M:%S")


            event = self.env['event.event'].search([('mrm_id','=',vals['mrm_id'])],limit=1)
            event.write(vals)

    def fix_iht_event_dates(self):
        attachment = self.env['ir.attachment'].browse(self.id)
        objFile = str(base64.b64decode(attachment.datas).decode('UTF-8'))
        reader = objFile.splitlines()
        for row in reader:
            lines = row.split('\t')
            vals = {}
            vals['mrm_id'] = lines[1]

            date_begin = lines[5].replace("/","-")
            date_begin = date_begin + " 00:00:00"

            date_end = lines[6].replace("/","-")
            date_end = date_end + " 00:00:00"

            #raise UserError(date_begin)
            vals['date_begin'] = datetime.strptime (date_begin, "%d-%m-%Y %H:%M:%S")
            vals['date_end'] = datetime.strptime (date_end, "%d-%m-%Y %H:%M:%S")

            event = self.env['event.event'].search([('mrm_id','=',vals['mrm_id'])],limit=1)
            event.write(vals)



    def import_europe_quotations(self):

        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()
        for row in reader:
            lines = row.split('\t')
            vals = {}

            vals['mrm'] = True

            if lines[0]:
                vals['name'] = 'SO ' + lines[0]

            if lines[1]:
                #date = datetime.strptime(lines[1], '%b %d, %Y')
                date = dateparser.parse(lines[1]).date()
                #raise UserError(date)
                vals['date_order'] = date

            if lines[3]:
                try:
                    partner = self.env['res.partner'].search([('tfrm_id','=',int(lines[3]))],limit=1)

                except:
                    raise UserError(lines)

                if not partner:
                    partner = self.env['res.partner'].search([('name','=',lines[2])],limit=1)

                    if not partner:
                        partner = self.env['res.partner'].create({
                            'name': lines[2],
                            'tfrm_id': lines[3],
                            'company_type': 'person',
                            'from_excel': True,
                            })

                    else:
                        partner.write({
                            'name': lines[2],
                            'tfrm_id': lines[3],
                            'company_type': 'person',
                            'from_excel': True,
                            })

                else:
                    partner.write({
                        'name': lines[2],
                        'tfrm_id': lines[3],
                        'company_type': 'person',
                        'from_excel': True,
                        })



                vals['partner_id'] = partner.id

            vals['note'] = lines[30]

            if lines[27]:
                company_name = lines[27].strip()
                company = self.env['res.company'].search([('name','=',company_name)],limit=1)
                if not company:
                    company = self.env['res.company'].search([('name','ilike',company_name)],limit=1)
                vals['company_id'] = company.id

                currency = self.env['res.currency'].search([('name','=',lines[17])])
                pricelist = self.env['product.pricelist'].search([('company_id','=',company.id),('currency_id','=',currency.id),('name','=',company.name)])
                if not pricelist:
                    pricelist = self.env['product.pricelist'].search([('currency_id','=',currency.id),('name','=',company.name)])
                    
                if not pricelist:
                    try:
                        pricelist = self.env['product.pricelist'].create({
                            'company_id': company.id,
                            'name': company.name,
                            'currency_id': currency.id,
                            })
                    except:
                        raise UserError(self.name + ' - ' + str(i) + ' - ' + lines[20])
                
                vals['pricelist_id'] = pricelist.id


            if 'date_order' in vals:
                order = self.env['sale.order'].search([('name','=',vals['name'])])
                if not order:
                    if 'partner_id' in vals:
                        order = self.env['sale.order'].create(vals)
                        order.write({'name': vals['name']})

        self.is_processed = True


    def import_europe_quotation_lines(self):
        count = self.quotation_lines_row_count
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        for row in reader:
            lines = row.split('\t')
            vals = {}
            date = False
            if lines[0]:
                order_name = 'SO ' + str(lines[0])
                order = self.env['sale.order'].search([('name','=',order_name)])

            if lines[1]:
                date = dateparser.parse(lines[1]).date()

                if order:
                    if lines[17]:
                        company = self.env['res.company'].search([('name','=',lines[27])],limit=1)

                        if lines[16]:
                            product = self.env['product.template'].search([('tfrm_id','=',int(lines[16]))],limit=1)
                            if not product:
                                product = self.env['product.template'].create({
                                    'tfrm_id': int(lines[16]),
                                    'name': lines[15],
                                    'type': 'product' if lines[29] == 'TRUE' else 'service'
                                    })
                            else:
                                product.write({
                                    'name': lines[16],
                                    'type': 'product' if lines[29] == 'TRUE' else 'service'
                                    })
                            
                            product_variant = product.product_variant_id
                            vals['order_id'] = order.id
                            vals['product_id'] = product_variant.id
                            vals['product_uom'] = product_variant.uom_id.id
                            vals['name'] = lines[15]
                            vals['product_uom_qty'] = float(lines[19]) if lines[19] else 0
                            vals['discount'] = float(lines[20]) if lines[20] else 0
                            vals['price_unit'] = float(lines[18]) if lines[18] else 0

                            tax = False
                            if lines[23] and float(lines[23]) > 0:
                                tax = self.env['account.tax'].search([('mrm_id','>',0),('company_id','=',company.id),('amount','=',float(lines[23]))])
                                if not tax:
                                    tax = self.env['account.tax'].search([('mrm_id','>',0),('company_id','=',company.parent_id.id),('amount','=',float(lines[23]))])
                                if not tax:
                                    tax = self.env['account.tax'].search([('company_id','=',company.parent_id.id),('amount','=',float(lines[23]))])


                                vals['tax_id'] = [(6, 0, [tax.id])] if tax else [(5, 0, 0)]

                            term = self.env['term'].search([('name','=',lines[26])],limit=1)
                            vals['term_id'] = term.id
                            if not term:
                                term = self.env['term'].create({
                                    'name': lines[26]
                                    })

                            if lines[28]:
                                event = self.env['event.event'].search([('mrm_id','=',int(lines[28]))])
                                vals['event_id'] = event.id
                            
                            order_line = self.env['sale.order.line'].create(vals)

                    order.action_confirm()
                    order.write({'date_order': date})

        self.lines_processed = True


    def import_europe_receipts(self):
        payment_types = {
        'Bank Transfer': 'transfer',
        'Cash': 'cash',
        'Cheque': 'cheque',
        'Converted FP (Void)': 'cash',
        'Credit/Debit Card': 'credit_card',
        }
        count = self.processed_rows_count
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        for row in reader:
            lines = row.split('\t')
            vals = {}
            vals['from_mrm'] = True
            vals['partner_type'] = 'customer'
            vals['payment_method_id'] = self.env.ref('account.account_payment_method_manual_in').id

            partner = self.env['res.partner'].search([('mrm_id','=',lines[1]),('is_mrm_contact','=',False)])
            vals['partner_id'] = partner.id
            vals['name'] = lines[2]


            payment_date = lines[3]
            payment_date = payment_date.replace("/", "-")
            payment_date = payment_date[:-2] + "20" + payment_date[-2:]
            payment_date = datetime.strptime(payment_date, "%d-%m-%Y")
            vals['payment_date'] = payment_date

            
            if lines[4]:
                date_cleared = lines[4]
                date_cleared = date_cleared.replace("/", "-")
                date_cleared = date_cleared[:-2] + "20" + date_cleared[-2:]
                date_cleared = datetime.strptime(date_cleared, "%d-%m-%Y")
                vals['date_cleared'] = date_cleared
            
            #vals['invoice_ids'] = [(6, 0, [invoice.id])]

            invoice = self.env['account.move'].search([('name','=',lines[5])])

            #company = self.env['res.company'].search([('name','=',lines[13].strip())])
            #if not company:
                #company = self.env['res.company'].search([('name','ilike',lines[13].strip())])
            
            if invoice:
                company = invoice.company_id
                vals['company_id'] = company.id


                
                vals['payment_mode'] = payment_types.get(lines[6])

                if vals['payment_mode'] == 'cash':
                    journal = self.env['account.journal'].search([('company_id','=',company.id),('type','=','cash')],limit=1)

                else:
                    journal = self.env['account.journal'].search([('company_id','=',company.id),('type','=','bank')],limit=1)

                vals['journal_id'] = journal.id

                vals['payment_ref'] = lines[7]

                currency = self.env['res.currency'].search([('name','=',lines[8])])
                vals['currency_id'] = currency.id

                amount = float(lines[9]) if lines[8] != 'USD' else float(lines[10])
                vals['amount'] = -amount if amount < 0 else amount

                vals['payment_type'] = 'inbound' if amount >= 0 else 'outbound'

                vals['comment'] = lines[11]

                payment = self.env['account.payment'].search([('name','=',vals['name']),('from_mrm','=',True)])
                if not payment:
                    payment = self.env['account.payment'].create(vals)
                    try:
                        payment.post()

                    except:
                        raise UserError(lines[2])

                elif payment.state == 'posted':
                    payment.action_draft()
                    amount = payment.amount
                    vals['amount'] += amount
                    payment.write(vals)
                    try:
                        payment.post()
                    except:
                        raise UserError(lines[2])


                register_payment = self.env['register.payment'].search([('payment_id','=',payment.id)])
                if not register_payment:
                    register_payment = self.env['register.payment'].create({
                        'payment_id': payment.id,
                        'partner_id': vals['partner_id'],
                        'from_mrm': True,
                        })


                payment_line = self.env['register.payment.line'].search([('move_id','=',invoice.id),('payment_registration_id','=',register_payment.id)])
                if not payment_line:
                    allocated_amount = float(lines[9])
                    allocated_amount = -allocated_amount if allocated_amount < 0 else allocated_amount
                    payment_line = self.env['register.payment.line'].create({
                        'payment_registration_id': register_payment.id,
                        'move_id': invoice.id,
                        'amount': allocated_amount,
                        'allocated_amount': allocated_amount,
                        'partner_id': vals['partner_id'],
                        })

        self.is_processed = True

    def import_all_instructors(self):

        payment_methods = {
        'Transfer': 'transfer',
        'Check': 'cheque',
        'Cash': 'cash',
        }

        marital = {
        'Divorced': 'divorced',
        'Single': 'single',
        'Widowed': 'widower',
        'Married': 'married',
        }

        countries = {
        'American': 'United States',
        'Austrailian': 'Australia',
        'Australia': 'Australia',
        'Canadian': 'Canada',
        'Canada': 'Canada',
        'Egypt': 'Egypt',
        'Egyptian': 'Egypt',
        'Egyption': 'Egypt',
        'Filipino': 'Philippines',
        'Indian': 'India',
        'Jordan': 'Jordan',
        'Jordanian': 'Jordan',
        'Lebanese': 'Lebanon',
        'Lebenese': 'Lebanon',
        'Lebanon': 'Lebanon',
        'Mexican': 'Mexico',
        'Pakastani': 'Pakistan',
        'Pakistan': 'Pakistan',
        'Pakistani': 'Pakistan',
        'Palestinian': 'State of Palestine',
        'Saudi': 'Saudi Arabia',
        'Syrian': 'Syria',
        'U.S Citizen': 'United States',
        'UAE': 'United Arab Emirates',
        'KSA': 'Saudi Arabia',
        'Luxembourg': 'Luxembourg',
        'Oman': 'Oman',
        'Philippines': 'Philippines',
        'Fillipino': 'Philippines',
        'Arab American': 'United States',
        'Juwaya': 'Lebanon',
        'Karachi, Pakistan': 'Pakistan',
        'Kashmir, India': 'India',
        'Serlion': 'serlion',
        'South Austrailia': 'Australia',
        'United Kingdom': 'United Kingdom',
        'Switzerland': 'Switzerland',
        'Saudi Arabia': 'Saudi Arabia',
        'Egyptain': 'Egypt',
        'Palestine': 'State of Palestine',
        'Saudia Arabia': 'Saudia Arabia',
        'Kuwait': 'Kuwait',
        'Lebenon': 'Lebanon',
        'Syria': 'Syria',
        'India': 'India',
        'Mexico': 'Mexico',
        'Eygpt': 'Eygpt',
        'lebanese': 'Lebanon'
        }

        gender = {
        'Male': 'male',
        'Female': 'female',
        }

        wb_obj = load_workbook(filename = 'data/filestore/morgan13-production-1206400/d6/Updated Active database.xlsx')
        wsheet = wb_obj['Sheet1']
        dataDict = {}


        counter = 0
        for key, *values in wsheet.iter_rows():
            key = str(counter)
            dataDict[key] = [v.value for v in values]
            counter = counter + 1


        for key, value in dataDict.items():
            vals = {}
            if value[0] and value[0] != '#REF!':
                vals['mrm_id'] = int(value[0])
                vals['name'] = value[1].strip()

                if value[2]:
                    company_name = value[2].rstrip()
                    company = self.env['res.company'].search([('name','=',company_name)])
                    if not company:
                        company = self.env['res.company'].search([('name','ilike',company_name)])
                    vals['company_id'] = company.id

                else:
                    vals['company_id'] = False

                if value[3]:
                    region_name = value[3].rstrip()
                    region = self.env['res.company'].search([('name','=',region_name)])
                    if not region:
                        region = self.env['res.company'].search([('name','ilike',region_name)])
                    
                    vals['recruiting_city_id'] = region.id if region else False
                    if region:
                        vals['company_id'] = region.parent_id.id
                
                
                vals['active'] = False if value[4] == 'TRUE' else True
                vals['mrm_status'] = 'dormant' if value[4] == 'TRUE' else 'active'
                vals['is_instructor'] = True

                vals['is_student'] = True if value[5] == 'Yes' else False

                if value[6]:
                    if isinstance(value[6], str):
                        #raise UserError(value[6])
                        try:
                            vals['joining_date'] = datetime.strptime(str(value[6]), "%d/%m/%Y")

                        except:
                            print("false")
                        
                    else:
                        vals['joining_date'] = value[6]


                vals['mobile_phone'] = value[7]
                
                vals['work_email'] = value[8]
                
                vals['work_location'] = value[9]
                    

                vals['private_email'] = value[14]
                
                mobile = str(value[15])
                if ".0" in mobile:
                    mobile = mobile.replace(".0","")

                vals['mobile'] = mobile
                
                bank = False
                if value[16]:
                    bank_name = value[16].strip()
                    bank = self.env['res.bank'].search([('name','=',bank_name)])
                    if not bank:
                        bank = self.env['res.bank'].create({'name': bank_name})

                
                partner_country = False
                if value[13]:
                    partner_country = self.env['res.country'].search([('name','=',countries[value[13].strip()])]).id

                partner = self.env['res.partner'].search([('name','=',vals['name'])],limit=1)
                if not partner:
                    partner = self.env['res.partner'].create({
                        'name': vals['name'],
                        'company_type': 'person',
                        'street': value[11],
                        'city': value[12],
                        'country_id': partner_country,
                        'mobile': vals['mobile'],
                        'email': vals['private_email'] if vals['private_email'] != 'NA' else False,
                        })

                else:
                    partner.write({
                        'name': vals['name'],
                        'company_type': 'person',
                        'street': value[11],
                        'city': value[12],
                        'country_id': partner_country,
                        'mobile': vals['mobile'],
                        'email': vals['private_email'] if vals['private_email'] != 'NA' else False,
                        })

                vals['address_home_id'] = partner.id

                if value[17]:
                    account_number = str(value[17]).strip()
                    if ".0" in account_number:
                        account_number = account_number.replace(".0","")
                    bank_account = self.env['res.partner.bank'].search([('acc_number','=',account_number)])
                    if not bank_account:
                        bank_account = self.env['res.partner.bank'].create({
                            'bank_id': bank.id if bank else False,
                            'acc_number': account_number,
                            'swift_code': value[18],
                            'iban': value[19],
                            'partner_id': partner.id,
                            'company_id': False
                            })


                vals['payment_method'] = payment_methods[value[20].strip()] if value[20] else False
                vals['marital'] = marital[value[21].strip()] if value[21] else False

                if value[22]:
                    children = str(value[22]).strip()
                    if children != '-' and children != 'None':
                        vals['children'] = int(children.split('.')[0])
                
                vals['emergency_contact'] = value[23]
                vals['emergency_phone'] = value[24]

                if value[25]:
                    country_name = value[25].strip()
                    if '/' in country_name:
                        country_array = country_name.split('/')
                        country_name = country_array[0]
                        vals['notes'] = 'Other Nationality: ' + country_array[1]
                    
                    odoo_country = countries[country_name]
                    vals['country_id'] = self.env['res.country'].search([('name','=',odoo_country)]).id

                vals['gender'] = gender[value[26].strip()] if value[26] else False

                if value[27]:
                    if isinstance(value[27], str):
                        #raise UserError(value[6])
                        try:
                            vals['birthday'] = datetime.strptime(str(value[27]), "%d/%m/%Y")

                        except:
                            print("false")
                        
                    else:
                        vals['birthday'] = value[27]

                if value[28]:
                    country_of_birth = value[28].strip()
                    vals['country_of_birth'] = self.env['res.country'].search([('name','=',countries[country_of_birth])]).id

                vals['social_twitter'] = value[29]
                
                biography = value[30]
                if biography:
                    biography = biography.replace("\n","<br/>")

                vals['biography'] = biography

                if value[31]:
                    category_name = value[31].strip()
                    category = self.env['hr.employee.category'].search([('name','=',category_name)])
                    if not category:
                        category = self.env['hr.employee.category'].create({
                            'name': category_name
                            })

                    vals['category_ids'] = [(6, 0, [category.id])]

                instructor = self.env['hr.employee'].search([('mrm_id','=',vals['mrm_id']),('active','in',[True,False])])
                if not instructor:
                    instructor = self.env['hr.employee'].create(vals)

                else:
                    try:
                        instructor.write(vals)

                    except:
                        vals['joining_date'] = False
                        try:
                            instructor.write(vals)

                        except:
                            vals['birthday'] = False
                            instructor.write(vals)

    def fix_invoice_lines(self):
        count = self.quotation_lines_row_count
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        new_value = count + 1000
        if new_value > len(reader):
            new_value = len(reader)

            

            self.lines_processed = True

        for i in range(count, new_value):
            lines = reader[i].split('\t')
            date = False
            if lines[0]:
                invoice_name = str(lines[0])


                invoice = self.env['account.move'].search([('name','=',invoice_name)])
                if not invoice:
                    invoice_vals = {}
                    invoice_vals['name'] = invoice_name
                    invoice_vals['type'] = 'out_refund' if 'CR' in invoice_name else 'out_invoice'


                    if lines[1]:
                        date = datetime.strptime(lines[1], '%b %d, %Y')
                        invoice_vals['invoice_date_due'] = date
                        invoice_vals['date'] = date
                        invoice_vals['invoice_date'] = date

                    if lines[3]:
                        try:
                            partner = self.env['res.partner'].search([('mrm_id','=',int(lines[3])),('is_mrm_contact','=',False)],limit=1)

                        except:
                            raise UserError(lines)

                        if not partner:
                            partner = self.env['res.partner'].search([('name','=',lines[2])],limit=1)

                            if not partner:
                                partner = self.env['res.partner'].create({
                                    'name': lines[2],
                                    'mrm_id': lines[3],
                                    'company_type': 'person',
                                    'from_excel': True,
                                    })

                            else:
                                partner.write({
                                    'name': lines[2],
                                    'mrm_id': lines[3],
                                    'company_type': 'person',
                                    'from_excel': True,
                                    })

                        else:
                            partner.write({
                                'name': lines[2],
                                'mrm_id': lines[3],
                                'company_type': 'person',
                                'from_excel': True,
                                })

                    else:
                        partner = self.env['res.partner'].browse(426764)

                    invoice_vals['partner_id'] = partner.id

                    invoice_vals['other_inv_ref'] = lines[4]
                    invoice_vals['note'] = lines[23]

                    if lines[20]:
                        company_name = lines[20].strip()
                        company = self.env['res.company'].search([('name','=',company_name)],limit=1)

                        if not company:
                            company = self.env['res.company'].search([('name','ilike',company_name)],limit=1)

                        invoice_vals['company_id'] = company.id

                    invoice_vals['journal_id'] = self.env['account.move'].with_context(force_company=company.id, default_type='out_invoice')._get_default_journal().id
                    currency = self.env['res.currency'].search([('name','=',lines[10])])
                    invoice_vals['currency_id'] = currency.id


                    if 'partner_id' in invoice_vals:
                        invoice = self.env['account.move'].create(invoice_vals)
                        invoice.write({'name': invoice_vals['name']})
                        invoice._onchange_invoice_line_ids()

                else:
                    invoice.button_draft()
                    invoice._onchange_invoice_line_ids()

                if lines[20]:
                    company = self.env['res.company'].search([('name','=',lines[20])],limit=1)

                    if not lines[9] and lines[12] and lines[0]:
                        if float(lines[11]) != 0:
                            invoice.button_draft()
                            if invoice.payment_ids:
                                payments = invoice.payment_ids.ids
                                self.env['register.payment'].search([('payment_id','in',payments)]).filtered(lambda p: p.is_processed == True).write({'is_processed': False})
                        

                        product = self.env['product.template'].sudo().browse(6247)

                    elif lines[9] and lines[0]:
                        product = self.env['product.template'].sudo().search([('mrm_id','=',int(lines[9]))],limit=1)
                        if not product:
                            product = self.env['product.template'].sudo().create({
                                'mrm_id': int(lines[9]),
                                'name': lines[8],
                                'type': 'product' if lines[22] == 'TRUE' else 'service'
                                })
                        else:
                            product.write({
                                'name': lines[8],
                                'type': 'product' if lines[22] == 'TRUE' else 'service'
                                })

                    
                    vals = {}
                    product_variant = product.product_variant_id
                    vals['move_id'] = invoice.id
                    vals['product_id'] = product_variant.id
                    vals['partner_id'] = invoice.partner_id.id
                    vals['product_uom_id'] = product_variant.uom_id.id
                    vals['name'] = lines[8] if lines[8] else product.name
                    vals['quantity'] = float(lines[12]) if lines[12] else 0
                    vals['discount'] = float(lines[13]) if lines[13] else 0

                    if vals['discount'] == 0 and lines[14] and float(lines[14]) > 0:
                        vals['discount'] = float(lines[14]) * 100 / float(lines[11]) if float(lines[11]) > 0 else 0 

                    price = 0
                    if lines[11]:
                        price = float(lines[11])
                        if price < 0 and invoice.type == 'out_refund':
                            price = -price
                    
                    vals['price_unit'] = price

                    vals['account_id'] = self.env['account.account'].sudo().search([('company_id','=', invoice.company_id.id),('code','=','400000'),('name','=','Product Sales')],limit=1).id
                    vals['exclude_from_invoice_tab'] = False


                    tax = False
                    if lines[16] and float(lines[16]) > 0:
                        tax = self.env['account.tax'].sudo().search([('mrm_id','>',0),('company_id','=',company.id),('amount','=',float(lines[16]))])
                        if not tax:
                            tax = self.env['account.tax'].sudo().search([('mrm_id','>',0),('company_id','=',company.parent_id.id),('amount','=',float(lines[16]))])

                        if not tax:
                            tax = self.env['account.tax'].sudo().search([('company_id','=',company.id),('amount','=',float(lines[16]))])

                        if not tax:
                            tax = self.env['account.tax'].sudo().search([('company_id','=',company.parent_id.id),('amount','=',float(lines[16]))])

                        if not tax:
                            tax = self.env['account.tax'].sudo().search([('company_id','=',invoice.company_id.id),('amount','=',float(lines[16]))])
                            if not tax:
                                tax_name = str(lines[16]) + '%'
                                if company:
                                    tax_name = company.name + ' - ' + str(lines[16]) + '%'

                                tax = self.env['account.tax'].create({
                                    'name': tax_name,
                                    'type_tax_use': 'sale',
                                    'company_id': company.id if company else invoice.company_id.id,
                                    'amount': float(lines[16]),
                                    'description': tax_name,
                                    })




                        vals['tax_ids'] = [(6, 0, [tax.id])] if tax else [(5, 0, 0)]

                    if lines[19]:
                        term = self.env['term'].search([('name','=',lines[19].strip())],limit=1)
                        if not term:
                            term = self.env['term'].create({
                                'name': lines[19].strip()
                                })
                        vals['term_id'] = term.id
                    
                    #create empty lines or unmigrated lines
                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                    if move_line and vals['product_id'] == 5850:
                        vals['product_id'] = 5851
                        move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                        if move_line:
                            vals['product_id'] = 5852
                            move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                            if move_line:
                                vals['product_id'] = 5853
                                move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5854
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5855
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5856
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5857
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5858
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5859
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)



                    if not move_line and vals['move_id']:
                        move_line = self.env['account.move.line'].sudo().create(vals)

                    #fix lines taxes

                    elif move_line and not move_line.tax_ids and tax:
                        invoice.button_draft()
                        if invoice.payment_ids:
                            payments = invoice.payment_ids.ids
                            self.env['register.payment'].search([('payment_id','in',payments)]).filtered(lambda p: p.is_processed == True).write({'is_processed': False})
                        
                        move_line.sudo().write({'tax_ids': vals['tax_ids'], 'line_updated': True})

                    #fix lines discount
                    elif move_line.discount == 0 and lines[14] and float(lines[14]) > 0:
                        invoice.button_draft()
                        if invoice.payment_ids:
                            payments = invoice.payment_ids.ids
                            self.env['register.payment'].search([('payment_id','in',payments)]).filtered(lambda p: p.is_processed == True).write({'is_processed': False})
                        discount = float(lines[14]) * 100 / float(lines[11]) if float(lines[11]) > 0 else 0 
                        move_line.sudo().write({'discount': discount, 'line_updated': True})

                    #fix lines without term
                    elif 'term_id' in vals and not move_line.term_id:
                        move_line.sudo().write({'term_id': vals['term_id'], 'line_updated': True}) 


        self.write({'quotation_lines_row_count': new_value})
        


    def fix_duplicated_empty_lines(self):
        count = self.quotation_lines_row_count
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        new_value = count + 5000
        if new_value > len(reader):
            new_value = len(reader)
            

            self.lines_processed = True

        for i in range(count, new_value):
            lines = reader[i].split('\t')
            date = False
            if lines[0]:
                invoice_name = str(lines[0])


                invoice = self.env['account.move'].search([('name','=',invoice_name)])
                if not invoice:
                    invoice_vals = {}
                    invoice_vals['name'] = invoice_name
                    invoice_vals['type'] = 'out_refund' if 'CR' in invoice_name else 'out_invoice'


                    if lines[1]:
                        date = datetime.strptime(lines[1], '%b %d, %Y')
                        invoice_vals['invoice_date_due'] = date
                        invoice_vals['date'] = date
                        invoice_vals['invoice_date'] = date

                    if lines[3]:
                        try:
                            partner = self.env['res.partner'].search([('mrm_id','=',int(lines[3]))],limit=1)

                        except:
                            raise UserError(lines)

                        if not partner:
                            partner = self.env['res.partner'].search([('name','=',lines[2])],limit=1)

                            if not partner:
                                partner = self.env['res.partner'].create({
                                    'name': lines[2],
                                    'mrm_id': lines[3],
                                    'company_type': 'person',
                                    'from_excel': True,
                                    })

                            else:
                                partner.write({
                                    'name': lines[2],
                                    'mrm_id': lines[3],
                                    'company_type': 'person',
                                    'from_excel': True,
                                    })

                        else:
                            partner.write({
                                'name': lines[2],
                                'mrm_id': lines[3],
                                'company_type': 'person',
                                'from_excel': True,
                                })

                        invoice_vals['partner_id'] = partner.id

                        invoice_vals['other_inv_ref'] = lines[4]
                        invoice_vals['note'] = lines[23]

                        if lines[20]:
                            company_name = lines[20].strip()
                            company = self.env['res.company'].search([('name','=',company_name)],limit=1)

                            if not company:
                                company = self.env['res.company'].search([('name','ilike',company_name)],limit=1)

                            invoice_vals['company_id'] = company.id

                        invoice_vals['journal_id'] = self.env['account.move'].with_context(force_company=company.id, default_type='out_invoice')._get_default_journal().id

                        currency = self.env['res.currency'].search([('name','=',lines[10])])
                        invoice_vals['currency_id'] = currency.id

                    if 'partner_id' in invoice_vals:
                        invoice = self.env['account.move'].create(invoice_vals)
                        invoice.write({'name': invoice_vals['name']})

                if lines[20]:
                    company = self.env['res.company'].search([('name','=',lines[20])],limit=1)

                    if not lines[9] and lines[12] and lines[0]:
                        if float(lines[11]) != 0:
                            invoice.button_draft()
                            if invoice.payment_ids:
                                payments = invoice.payment_ids.ids
                                self.env['register.payment'].search([('payment_id','in',payments)]).filtered(lambda p: p.is_processed == True).write({'is_processed': False})
                        

                        product = self.env['product.template'].sudo().browse(6247)
                    
                        vals = {}
                        product_variant = product.product_variant_id
                        vals['move_id'] = invoice.id
                        vals['product_id'] = product_variant.id
                        vals['partner_id'] = invoice.partner_id.id
                        vals['product_uom_id'] = product_variant.uom_id.id
                        vals['name'] = lines[8] if lines[8] else product.name
                        vals['quantity'] = float(lines[12]) if lines[12] else 0
                        vals['discount'] = float(lines[13]) if lines[13] else 0

                        price = 0
                        if lines[11]:
                            price = float(lines[11])
                            if price < 0 and invoice.type == 'out_refund':
                                price = -price
                        
                        vals['price_unit'] = price

                        vals['account_id'] = self.env['account.account'].sudo().search([('company_id','=', invoice.company_id.id),('code','=','400000'),('name','=','Product Sales')],limit=1).id
                        vals['exclude_from_invoice_tab'] = False


                        tax = False
                        if lines[16] and float(lines[16]) > 0:
                            tax = self.env['account.tax'].sudo().search([('mrm_id','>',0),('company_id','=',company.id),('amount','=',float(lines[16]))])
                            if not tax:
                                tax = self.env['account.tax'].sudo().search([('mrm_id','>',0),('company_id','=',company.parent_id.id),('amount','=',float(lines[16]))])

                            if not tax:
                                tax = self.env['account.tax'].sudo().search([('company_id','=',company.id),('amount','=',float(lines[16]))])

                            if not tax:
                                tax = self.env['account.tax'].sudo().search([('company_id','=',company.parent_id.id),('amount','=',float(lines[16]))])



                            vals['tax_ids'] = [(6, 0, [tax.id])] if tax else [(5, 0, 0)]

                        if lines[19]:
                            term = self.env['term'].search([('name','=',lines[19].strip())],limit=1)
                            if not term:
                                term = self.env['term'].create({
                                    'name': lines[19].strip()
                                    })
                            vals['term_id'] = term.id
                    
                        #create empty lines or unmigrated lines
                        move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False)],limit=1)
                        if move_line and vals['product_id'] == 5850:
                            vals['product_id'] = 5851
                            move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False)],limit=1)
                            if move_line:
                                vals['product_id'] = 5852
                                move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5853
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False)],limit=1)
                                    if move_line:
                                        vals['product_id'] = 5854
                                        move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False)],limit=1)



                        if not move_line:
                            move_line = self.env['account.move.line'].sudo().create(vals)


        self.write({'quotation_lines_row_count': new_value})

    def fix_invoice_europe(self):
        count = self.quotation_lines_row_count
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        new_value = count + 5000
        if new_value > len(reader):
            new_value = len(reader)

            

            self.lines_processed = True

        for i in range(count, new_value):
            lines = reader[i].split('\t')
            date = False
            if lines[0]:
                invoice_name = str(lines[0])


                invoice = self.env['account.move'].search([('name','=',invoice_name)])
                if invoice:
                    invoice.button_draft()
                
                else:
                    invoice_vals = {}
                    invoice_vals['name'] = invoice_name
                    invoice_vals['type'] = 'out_refund' if 'CR' in invoice_name else 'out_invoice'


                    if lines[1]:
                        date = dateparser.parse(lines[1]).date()
                        invoice_vals['invoice_date_due'] = date
                        invoice_vals['date'] = date
                        invoice_vals['invoice_date'] = date

                    if lines[3]:
                        try:
                            partner = self.env['res.partner'].search([('tfrm_id','=',int(lines[3]))],limit=1)

                        except:
                            raise UserError(lines)

                        if not partner:
                            partner = self.env['res.partner'].search([('name','=',lines[2])],limit=1)

                            if not partner:
                                partner = self.env['res.partner'].create({
                                    'name': lines[2],
                                    'tfrm_id': lines[3],
                                    'company_type': 'person',
                                    'from_excel': True,
                                    })

                            else:
                                partner.write({
                                    'name': lines[2],
                                    'tfrm_id': lines[3],
                                    'company_type': 'person',
                                    'from_excel': True,
                                    })

                        else:
                            partner.write({
                                'name': lines[2],
                                'tfrm_id': lines[3],
                                'company_type': 'person',
                                'from_excel': True,
                                })

                        invoice_vals['partner_id'] = partner.id

                        invoice_vals['note'] = lines[30]

                        if lines[27]:
                            company_name = lines[27].strip()
                            company = self.env['res.company'].search([('name','=',company_name)],limit=1)

                            if not company:
                                company = self.env['res.company'].search([('name','ilike',company_name)],limit=1)

                            invoice_vals['company_id'] = company.id

                        invoice_vals['journal_id'] = self.env['account.move'].with_context(force_company=company.id, default_type='out_invoice')._get_default_journal().id

                        currency = self.env['res.currency'].search([('name','=',lines[17])])
                        invoice_vals['currency_id'] = currency.id

                    if 'partner_id' not in invoice_vals:
                        invoice_vals['partner_id'] = 115122
                    
                    invoice = self.env['account.move'].create(invoice_vals)
                    invoice.write({'name': invoice_vals['name']})


        self.write({'quotation_lines_row_count': new_value})


    def import_europe_missing_lines(self):
        count = self.quotation_lines_row_count
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        new_value = count + 5000
        if new_value > len(reader):
            new_value = len(reader)
            

            

            self.lines_processed = True

        for i in range(count, new_value):
            lines = reader[i].split('\t')
            date = False
            if lines[0]:
                invoice_name = str(lines[0])


                invoice = self.env['account.move'].search([('name','=',invoice_name)])
                if lines[27]:
                    company = self.env['res.company'].search([('name','=',lines[27].strip())],limit=1)

                    product = self.env['product.template'].browse(6247)
                    if lines[16]:
                        product = self.env['product.template'].search([('tfrm_id','=',int(lines[16]))],limit=1)
                        if not product:
                            product = self.env['product.template'].create({
                                'tfrm_id': int(lines[16]),
                                'name': lines[15],
                                'type': 'product' if lines[29] == 'TRUE' else 'service'
                                })

                        else:
                            product.write({
                                'name': lines[15],
                                'type': 'product' if lines[29] == 'TRUE' else 'service'
                                })
                    
                    vals = {}
                    product_variant = product.product_variant_id
                    vals['move_id'] = invoice.id
                    vals['product_id'] = product_variant.id
                    vals['partner_id'] = invoice.partner_id.id
                    vals['product_uom_id'] = product_variant.uom_id.id
                    vals['name'] = lines[15] if lines[15] else product.name
                    vals['quantity'] = float(lines[19]) if lines[19] else 0
                    vals['discount'] = float(lines[20]) if lines[20] else 0

                    price = 0
                    if lines[18]:
                        price = float(lines[18])
                        if price < 0 and invoice.type == 'out_refund':
                            price = -price
                    
                    vals['price_unit'] = price

                    vals['account_id'] = self.env['account.account'].sudo().search([('company_id','=', invoice.company_id.id),('code','=','400000'),('name','=','Product Sales')],limit=1).id
                    vals['exclude_from_invoice_tab'] = False


                    tax = False
                    if lines[23] and float(lines[23]) > 0:
                        tax = self.env['account.tax'].search([('mrm_id','>',0),('company_id','=',company.id),('amount','=',float(lines[23]))])
                        if not tax:
                            tax = self.env['account.tax'].search([('mrm_id','>',0),('company_id','=',company.parent_id.id),('amount','=',float(lines[23]))])
                        if not tax:
                            tax = self.env['account.tax'].search([('company_id','=',company.parent_id.id),('amount','=',float(lines[23]))])



                        vals['tax_ids'] = [(6, 0, [tax.id])] if tax else [(5, 0, 0)]

                    if lines[26]:
                        term = self.env['term'].search([('name','=',lines[26].strip())],limit=1)
                        if not term:
                            term = self.env['term'].create({
                                'name': lines[26].strip()
                                })
                        vals['term_id'] = term.id

                    if vals['move_id']:
                        move_line = self.env['account.move.line'].sudo().create(vals)


    def import_europe_missing_lines(self):
        count = self.quotation_lines_row_count
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()
        for row in reader:
            lines = row.split('\t')
            receipt_name = lines[1].split()
            receipt = self.env['account.payment'].search([('name','=',receipt_name)])
            self.env['register.payment'].search([('payment_id','=',receipt.id)]).filtered(lambda p: p.is_processed == True).write({'is_processed': False})
            receipt.action_draft()
            receipt.post()

    def fix_stock_invoice_lines(self):
        count = self.quotation_lines_row_count
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        new_value = count + 5000
        if new_value > len(reader):
            new_value = len(reader)

            

            self.lines_processed = True

        for i in range(count, new_value):
            lines = reader[i].split('\t')
            date = False
            if lines[0]:
                invoice_name = str(lines[0])


                invoice = self.env['account.move'].search([('name','=',invoice_name)])
                if not invoice:
                    invoice_vals = {}
                    invoice_vals['name'] = invoice_name
                    invoice_vals['type'] = 'out_refund' if 'CR' in invoice_name else 'out_invoice'


                    if lines[1]:
                        date = datetime.strptime(lines[1], '%b %d, %Y')
                        invoice_vals['invoice_date_due'] = date
                        invoice_vals['date'] = date
                        invoice_vals['invoice_date'] = date

                    if lines[3]:
                        try:
                            partner = self.env['res.partner'].search([('mrm_id','=',int(lines[3])),('is_mrm_contact','=',False)],limit=1)

                        except:
                            raise UserError(lines)

                        if not partner:
                            partner = self.env['res.partner'].search([('name','=',lines[2]),('is_mrm_contact','=',False)],limit=1)

                            if not partner:
                                partner = self.env['res.partner'].create({
                                    'name': lines[2],
                                    'mrm_id': lines[3],
                                    'company_type': 'person',
                                    'from_excel': True,
                                    })

                            else:
                                partner.write({
                                    'name': lines[2],
                                    'mrm_id': lines[3],
                                    'company_type': 'person',
                                    'from_excel': True,
                                    })

                        else:
                            partner.write({
                                'name': lines[2],
                                'mrm_id': lines[3],
                                'company_type': 'person',
                                'from_excel': True,
                                })

                        invoice_vals['partner_id'] = partner.id

                        invoice_vals['other_inv_ref'] = lines[4]
                        invoice_vals['note'] = lines[23]

                        if lines[20]:
                            company_name = lines[20].strip()
                            company = self.env['res.company'].search([('name','=',company_name)],limit=1)

                            if not company:
                                company = self.env['res.company'].search([('name','ilike',company_name)],limit=1)

                            invoice_vals['company_id'] = company.id

                        invoice_vals['journal_id'] = self.env['account.move'].with_context(force_company=company.id, default_type='out_invoice')._get_default_journal().id

                        currency = self.env['res.currency'].search([('name','=',lines[10])])
                        invoice_vals['currency_id'] = currency.id

                    if 'partner_id' in invoice_vals:
                        invoice = self.env['account.move'].create(invoice_vals)
                        invoice.write({'name': invoice_vals['name']})

                if lines[20]:
                    company = self.env['res.company'].search([('name','=',lines[20])],limit=1)

                    if not lines[9] and lines[12] and lines[0]:
                        product = self.env['product.template'].sudo().browse(6247)

                    elif lines[9] and lines[0]:
                        product = self.env['product.template'].sudo().search([('mrm_id','=',int(lines[9]))],limit=1)
                    
                    vals = {}
                    product_variant = product.product_variant_id
                    vals['move_id'] = invoice.id
                    vals['product_id'] = product_variant.id
                    vals['partner_id'] = invoice.partner_id.id
                    vals['product_uom_id'] = product_variant.uom_id.id
                    vals['name'] = lines[8] if lines[8] else product.name
                    vals['quantity'] = float(lines[12]) if lines[12] else 0
                    vals['discount'] = float(lines[13]) if lines[13] else 0

                    price = 0
                    if lines[11]:
                        price = float(lines[11])
                        if price < 0 and invoice.type == 'out_refund':
                            price = -price
                    
                    vals['price_unit'] = price

                    vals['account_id'] = self.env['account.account'].sudo().search([('company_id','=', invoice.company_id.id),('code','=','400000'),('name','=','Product Sales')],limit=1).id
                    vals['exclude_from_invoice_tab'] = False


                    tax = False
                    if lines[16] and float(lines[16]) > 0:
                        tax = self.env['account.tax'].sudo().search([('mrm_id','>',0),('company_id','=',company.id),('amount','=',float(lines[16]))])
                        if not tax:
                            tax = self.env['account.tax'].sudo().search([('mrm_id','>',0),('company_id','=',company.parent_id.id),('amount','=',float(lines[16]))])

                        if not tax:
                            tax = self.env['account.tax'].sudo().search([('company_id','=',company.id),('amount','=',float(lines[16]))])

                        if not tax:
                            tax = self.env['account.tax'].sudo().search([('company_id','=',company.parent_id.id),('amount','=',float(lines[16]))])

                        if not tax:
                            tax = self.env['account.tax'].sudo().search([('company_id','=',invoice.company_id.id),('amount','=',float(lines[16]))])
                            if not tax:
                                tax_name = str(lines[16]) + '%'
                                if company:
                                    tax_name = company.name + ' - ' + str(lines[16]) + '%'

                                tax = self.env['account.tax'].create({
                                    'name': tax_name,
                                    'type_tax_use': 'sale',
                                    'company_id': company.id if company else invoice.company_id.id,
                                    'amount': float(lines[16]),
                                    'description': tax_name,
                                    })




                        vals['tax_ids'] = [(6, 0, [tax.id])] if tax else [(5, 0, 0)]

                    if lines[19]:
                        term = self.env['term'].search([('name','=',lines[19].strip())],limit=1)
                        if not term:
                            term = self.env['term'].create({
                                'name': lines[19].strip()
                                })
                        vals['term_id'] = term.id
                    
                    #create empty lines or unmigrated lines
                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                    if move_line and vals['product_id'] == 5850:
                        vals['product_id'] = 5851
                        move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                        if move_line:
                            vals['product_id'] = 5852
                            move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                            if move_line:
                                vals['product_id'] = 5853
                                move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5854
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5855
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5856
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5857
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5858
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5859
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)

                    if move_line:
                        if lines[24]:
                            move_line.write({'mrm_stock_id': lines[24]})

                        if lines[25]:
                            move_line.write({'mrm_stock_type': lines[25]})


        self.write({'quotation_lines_row_count': new_value})


    def fix_opportunities_issue(self):
        offset = self.env['ir.config_parameter'].browse(48)
        value = offset.value
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        study_format = {
        'In-House Training': 'inhouse',
        'Live Class': 'live',
        'Live Class, In-House Training, Workshop': 'multiple',
        'Live Class, Live Online': 'multiple',
        'Live Class, Self-Study (packages that help candidates pass)': 'multiple',
        'Live Class, Self-Study (packages that help candidates pass), Stand-Alone (individual products)': 'multiple',
        'Live Class, Self-Study (packages that help candidates pass), Live Online': 'multiple',
        'Live Online': 'liveonline',
        'Other': 'other',
        'Self-Study (packages that help candidates pass)': 'self',
        'Self-Study (packages that help candidates pass), Live Online': 'self',
        'Stand-Alone (individual products)': 'standalone',
        'Workshop': 'workshop',
        'Workshop, Other': 'workshop',
        'Workshops': 'workshop',
        'In-House Training, Live Online': 'inhouse',
        'Live Class, Workshop': 'live',
        'Live Class, Other': 'live',
        'Self-Study (packages that help candidates pass), Other': 'self',
        'Other, Live Online': 'liveonline',
        'Stand-Alone (individual products), In-House Training, Workshop': 'standalone',
        'Live Class, Workshop, Other': 'live',
        'Live Class, Self-Study (packages that help candidates pass), Other': 'live',
        'Live Class, In-House Training': 'live',
        'In-House Training, Workshop': 'inhouse'
        }


        stage = {
        'Submissions': 1,
        'Enquiries': 1,
        '1. Opportunity Initiation': 1,
        '2. General Inquiry': 1,
        '3. A - Course Inquiry': 2,
        '3. B - Interest unavailable': 10,
        '3. C - Marketing Event': 2,
        '4. A - Qualification': 3,
        '4. B - Nurturing Cycle': 3,
        '5. A - Prospecting': 3,
        '5. B - Non-Qualified': 10,
        '6. A - Current Candidate': 7,
        '6. B - Archived': 9,
        '6. C - Dead': 10,
        '7. A - Drop-Out': 10,
        '7. B - Alumni': 8,
        '7. C - Duplicate Opportunity': 10,
        }

        gender = {
        'Male': 'male',
        'Female': 'female',
        }

        new_value = int(value) + 2000
        if new_value > len(reader):
            new_value = len(reader)

        for i in range(int(value), new_value):

            lines = reader[i].split('\t')
            vals = {}




            vals['mrm_id'] = int(lines[0])

            user = self.env['res.users'].sudo().search([('name','=',lines[11].strip())],limit=1)
            if not user:
                user = self.env['res.users'].sudo().search([('name','ilike',lines[11].strip())],limit=1)
            
            if user:
                vals['user_id'] = user.id

            if lines[15]:
                vals['study_format'] = study_format[lines[15].strip()]

            if lines[17]:
                vals['probability'] = float(lines[17].replace("%",""))

            else:
                vals['probability'] = 0

            if lines[18] == 'Lost':
                vals['stage_id'] = 10

            if lines[19] != 'Website Order' and lines[19] != 'Website Order - Offline Payment':
                vals['stage_id'] = stage[lines[19].strip()] if 'stage_id' not in vals else 10

            
            vals['type'] = 'opportunity'
            vals['migration_type'] = 'opportunity'

            opportunity = self.env['crm.lead'].search([('mrm_id','=',vals['mrm_id']),('migration_type','=',vals['migration_type'])])
            opportunity.write(vals)

        offset.write({'value': str(new_value)})

    def import_contacts(self):
        count = self.processed_rows_count
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        gender = {
        'Male': 'male',
        'Female': 'female',
        }

        new_value = count + 1000
        if new_value > len(reader):
            new_value = len(reader)

            self.is_processed = True

        for i in range(count, new_value):
            lines = reader[i].split('\t')
            contact_vals = {}

            account = False
            account_vals = {}
            if lines[2]:
                try:
                    account_vals['mrm_id'] = int(lines[2].strip())
                    account_vals['name'] = lines[3].strip()
                    account = self.env['res.partner'].search([('mrm_id','=',account_vals['mrm_id']),('is_mrm_contact','=',False)],limit=1)
                    if not account:
                        account = self.env['res.partner'].create(account_vals)

                    else:
                        account.write(account_vals)

                except:
                    account = False

            contact_vals['mrm_id'] = int(lines[0].strip())
            contact_vals['name'] = lines[1].strip()
            contact_vals['is_mrm_contact'] = True
            contact_vals['parent_id'] = account.id if account else False
            contact_vals['company_type'] = 'person'
            contact_vals['account_type'] = 'b2c'

            country = False
            if lines[5]:
                country = self.env['res.country'].search([('name','=',lines[5].strip())])
                if not country:
                    country = self.env['res.country'].search([('name','ilike',lines[5].strip())],limit=1)

            if country:
                contact_vals['country_id'] = country.id

            if lines[6]:
                contact_vals['gender'] = gender.get(lines[6].strip())

            contact_vals['mobile'] = lines[7]
            contact_vals['phone'] = lines[8]
            contact_vals['email'] = lines[9]
            contact_vals['job_role'] = lines[10]
            contact_vals['function'] = lines[11]

            if lines[12]:
                nationality = self.env['res.country'].search([('code','=',lines[12].strip())])
                contact_vals['nationality'] = nationality.id

            contact_vals['university'] = lines[13].strip()
            contact_vals['city'] = lines[15].strip()
            contact_vals['street'] = lines[16].strip()
            contact_vals['zip'] = lines[17].strip()
            contact_vals['company_legal_name'] = lines[18].strip()
            contact_vals['vat'] = lines[19].strip()
            
            if lines[20]:
                education = self.env['education.level'].search([('name','=',lines[20].strip())])
                if not education:
                    education = self.env['education.level'].create({
                        'name': lines[20]
                        })

                contact_vals['education_level'] = education.id

            contact = self.env['res.partner'].search([('mrm_id','=',contact_vals['mrm_id']),('is_mrm_contact','=',True)])
            if not contact:
                contact = self.env['res.partner'].create(contact_vals)

            else:
                contact.write(contact_vals)




        self.write({'processed_rows_count': new_value})


    def archive_courses(self):
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()
        courses = []
        for row in reader:
            lines = row.split('\t')
            courses.append(int(lines[3]))

        courses_to_stay = self.env['course'].browse(courses)
        courses_to_archive = self.env['course'].search([]) - courses_to_stay
        
        programs_to_stay = courses_to_stay.mapped('program_id')
        programs_to_archive = self.env['program'].search([]) - programs_to_stay
        
        for program in programs_to_archive:
            program.write({'active': False})

        for course in courses_to_archive:
            course.write({'active': False})

    def archive_uneeded_users(self):
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()
        users = []
        for row in reader:
            lines = row.split('\t')
            user = self.env['res.users'].search([('login','=',lines[2])])
            users.append(user.id)

        users.append(2)

        users_to_stay = self.env['res.users'].browse(users)
        users_to_archive = self.env['res.users'].search([]) - users_to_stay

        for user in users_to_archive:
            user.write({'active': False})

    def import_future_payments(self):
        count = self.processed_rows_count
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()
        new_value = count + 500
        if new_value > len(reader):
            new_value = len(reader)
            self.is_processed = True

        
        for i in range(count, new_value):
            vals = {}
            lines = reader[i].split('\t')
            vals['from_mrm'] = True
            partner = self.env['res.partner'].search([('mrm_id','=',lines[1]),('is_mrm_contact','=',False)])

            invoice = self.env['account.move'].search([('name','=',lines[5])])
            vals['partner_id'] = invoice.partner_id.id
            vals['name'] = lines[2]
            vals['due_date'] = lines[3]
            vals['invoice_id'] = invoice.id

            vals['amount'] = float(lines[9])
            vals['comment'] = lines[11]
            vals['region_id'] = invoice.company_id.id
            vals['amount_usd'] = float(lines[10])

            term = self.env['term'].search([('name','=',lines[12].strip())],limit=1)
            if not term:
                term = self.env['term'].search([('name','ilike',lines[12].strip())],limit=1)

            vals['term_id'] = term.id

            future_payment = self.env['morgan.future.payment'].search([('invoice_id','=',vals['invoice_id']),('partner_id','=',vals['partner_id']),('name','=',vals['name'])])
            if not future_payment:
                future_payment = self.env['morgan.future.payment'].create(vals)
            else:
                future_payment.write(vals)

            future_payment.on_change_amount()

        self.write({'processed_rows_count': new_value})

    def import_mrm_future_tasks(self):
        count = self.processed_rows_count
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        new_value = count + 1000
        if new_value > len(reader):
            new_value = len(reader)

            self.is_processed = True

        for i in range(count, new_value):
            lines = reader[i].split('\t')
            if lines[1]:
                vals = {}
                vals['mrm_id'] = int(lines[0])

                lead = self.env['crm.lead'].search([('mrm_id','=',int(lines[1]))],limit=1)
                vals['res_id'] = lead.id

                user = self.env['res.users'].sudo().search([('name','=',lines[2].strip()),('active','in',[True,False])],limit=1)
                if not user:
                    user = self.env['res.users'].sudo().search([('name','ilike',lines[2].strip()),('active','in',[True,False])],limit=1)

                if not user:
                    user = lead.user_id

                if lead.company_id:
                    companies = user.company_ids.ids + lead.company_id.ids
                    user.write({'company_ids': [(6, 0, companies)]})

                vals['user_id'] = user.id
                vals['res_model'] = 'crm.lead'
                vals['res_model_id'] = self.env['ir.model'].sudo().search([('model','=','crm.lead')]).id

                activity_type = self.env['mail.activity.type'].sudo().search([('name','=',lines[4].strip())])

                if not activity_type:
                    activity_type = self.env['mail.activity.type'].sudo().create({
                        'name': lines[4].strip(),
                        'res_model_id': self.env['ir.model'].sudo().search([('model','=','crm.lead')]).id,
                        })

                vals['activity_type_id'] = activity_type.id
                vals['summary'] = lines[5]
                vals['note'] = lines[6]

                if isinstance(lines[7], str):
                    try:
                        vals['date_deadline'] = datetime.strptime(str(lines[7]), "%m/%d/%Y")

                    except:
                        raise UserError(lines)

                
                else:
                    vals['date_deadline'] = lines[7]

                activity = self.env['mail.activity'].sudo().search([('mrm_id','=',vals['mrm_id'])])
                if not activity:
                    if vals['user_id']:
                        try:
                            activity = self.env['mail.activity'].with_user(SUPERUSER_ID).create(vals)

                        except:
                            continue

                else:
                    activity.write(vals)

        self.write({'processed_rows_count': new_value})

    def fix_receipts_naming(self):
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()
        courses = []
        for row in reader:
            lines = row.split('\t')
            receipt = self.env['account.payment'].search([('name','=',lines[1]),('partner_id','=',False)])
            partner = self.env['res.partner'].search([('name','=',lines[4].strip())],limit=1)
            if partner:
                receipt.write({'partner_id': partner.id})
                items = self.env['account.move.line'].search([('move_id.name','=',receipt.name)])
                for item in items:
                    item.write({'partner_id': partner.id})

    def import_missing_invoice_lines(self):
        count = self.quotation_lines_row_count
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        new_value = count + 1000
        if new_value > len(reader):
            new_value = len(reader)

            

            self.lines_processed = True

        for i in range(count, new_value):
            lines = reader[i].split('\t')
            date = False
            if lines[0]:
                invoice_name = str(lines[0])


                invoice = self.env['account.move'].search([('name','=',invoice_name)])
                if not invoice:
                    invoice_vals = {}
                    invoice_vals['name'] = invoice_name
                    invoice_vals['type'] = 'out_refund' if 'CR' in invoice_name else 'out_invoice'


                    if lines[1]:
                        date = datetime.strptime(lines[1], '%b %d, %Y')
                        invoice_vals['invoice_date_due'] = date
                        invoice_vals['date'] = date
                        invoice_vals['invoice_date'] = date

                    if lines[3]:
                        try:
                            partner = self.env['res.partner'].search([('mrm_id','=',int(lines[3])),('is_mrm_contact','=',False)],limit=1)

                        except:
                            raise UserError(lines)

                        if not partner:
                            partner = self.env['res.partner'].search([('name','=',lines[2])],limit=1)

                            if not partner:
                                partner = self.env['res.partner'].create({
                                    'name': lines[2],
                                    'mrm_id': lines[3],
                                    'company_type': 'person',
                                    'from_excel': True,
                                    })

                            else:
                                partner.write({
                                    'name': lines[2],
                                    'mrm_id': lines[3],
                                    'company_type': 'person',
                                    'from_excel': True,
                                    })

                        else:
                            partner.write({
                                'name': lines[2],
                                'mrm_id': lines[3],
                                'company_type': 'person',
                                'from_excel': True,
                                })
                    else:
                        partner = self.env['res.partner'].browse(426764)

                    invoice_vals['partner_id'] = partner.id

                    invoice_vals['other_inv_ref'] = lines[4]
                    invoice_vals['note'] = lines[23]


                    if lines[20]:
                        company_name = lines[20].strip()
                        company = self.env['res.company'].search([('name','=',company_name)],limit=1)

                        if not company:
                            company = self.env['res.company'].search([('name','ilike',company_name)],limit=1)

                        invoice_vals['company_id'] = company.id

                        try:
                            invoice_vals['journal_id'] = self.env['account.move'].with_context(force_company=company.id, default_type='out_invoice')._get_default_journal().id

                        except:
                            raise UserError(i)

                    
                    currency = self.env['res.currency'].search([('name','=',lines[10])])
                    invoice_vals['currency_id'] = currency.id


                    if 'partner_id' in invoice_vals and invoice_vals['currency_id']:
                        invoice = self.env['account.move'].create(invoice_vals)
                        invoice.write({'name': invoice_vals['name']})

                else:
                    invoice.button_draft()
                    if invoice.payment_ids:
                        payments = invoice.payment_ids.ids
                        self.env['register.payment'].search([('payment_id','in',payments)]).filtered(lambda p: p.is_processed == True).write({'is_processed': False})

                if lines[20]:
                    company = self.env['res.company'].search([('name','=',lines[20])],limit=1)

                    if not lines[9] and lines[12] and lines[0]:
                        if float(lines[11]) != 0:
                            invoice.button_draft()
                            if invoice.payment_ids:
                                payments = invoice.payment_ids.ids
                                self.env['register.payment'].search([('payment_id','in',payments)]).filtered(lambda p: p.is_processed == True).write({'is_processed': False})
                        

                        product = self.env['product.template'].sudo().browse(6247)

                    elif lines[9] and lines[0]:
                        product = self.env['product.template'].sudo().search([('mrm_id','=',int(lines[9]))],limit=1)
                        if not product:
                            product = self.env['product.template'].sudo().create({
                                'mrm_id': int(lines[9]),
                                'name': lines[8],
                                'type': 'product' if lines[22] == 'TRUE' else 'service'
                                })
                        else:
                            product.write({
                                'name': lines[8],
                                'type': 'product' if lines[22] == 'TRUE' else 'service'
                                })
                    
                    vals = {}
                    product_variant = product.product_variant_id
                    vals['move_id'] = invoice.id
                    vals['product_id'] = product_variant.id
                    vals['partner_id'] = invoice.partner_id.id
                    vals['product_uom_id'] = product_variant.uom_id.id
                    vals['name'] = lines[8] if lines[8] else product.name
                    vals['quantity'] = float(lines[12]) if lines[12] else 0
                    vals['discount'] = float(lines[13]) if lines[13] else 0

                    if vals['discount'] == 0 and lines[14] and float(lines[14]) > 0:
                        vals['discount'] = float(lines[14]) * 100 / float(lines[11]) if float(lines[11]) > 0 else 0 

                    price = 0
                    if lines[11]:
                        price = float(lines[11])
                        if invoice.type == 'out_refund':
                            price = -price
                    
                    vals['price_unit'] = price

                    vals['account_id'] = self.env['account.account'].sudo().search([('company_id','=', invoice.company_id.id),('code','=','400000'),('name','=','Product Sales')],limit=1).id
                    vals['exclude_from_invoice_tab'] = False


                    tax = False
                    if lines[16] and float(lines[16]) > 0:
                        tax = self.env['account.tax'].sudo().search([('mrm_id','>',0),('company_id','=',company.id),('amount','=',float(lines[16]))])
                        if not tax:
                            tax = self.env['account.tax'].sudo().search([('mrm_id','>',0),('company_id','=',company.parent_id.id),('amount','=',float(lines[16]))])

                        if not tax:
                            tax = self.env['account.tax'].sudo().search([('company_id','=',company.id),('amount','=',float(lines[16]))])

                        if not tax:
                            tax = self.env['account.tax'].sudo().search([('company_id','=',company.parent_id.id),('amount','=',float(lines[16]))])

                        if not tax:
                            tax = self.env['account.tax'].sudo().search([('company_id','=',invoice.company_id.id),('amount','=',float(lines[16]))])
                            if not tax:
                                tax_name = str(lines[16]) + '%'
                                if company:
                                    tax_name = company.name + ' - ' + str(lines[16]) + '%'

                                tax = self.env['account.tax'].create({
                                    'name': tax_name,
                                    'type_tax_use': 'sale',
                                    'company_id': company.id if company else invoice.company_id.id,
                                    'amount': float(lines[16]),
                                    'description': tax_name,
                                    })




                        vals['tax_ids'] = [(6, 0, [tax.id])] if tax else [(5, 0, 0)]

                    if lines[19]:
                        term = self.env['term'].search([('name','=',lines[19].strip())],limit=1)
                        if not term:
                            term = self.env['term'].create({
                                'name': lines[19].strip()
                                })
                        vals['term_id'] = term.id
                    
                    #create empty lines or unmigrated lines
                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                    if move_line and vals['product_id'] == 5850:
                        vals['product_id'] = 5851
                        move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                        if move_line:
                            vals['product_id'] = 5852
                            move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                            if move_line:
                                vals['product_id'] = 5853
                                move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5854
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5855
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5856
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5857
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5858
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5859
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)



                    
                    move_line = self.env['account.move.line'].sudo().create(vals)


        self.write({'quotation_lines_row_count': new_value})

    def allocate_receipts(self):
        count = self.quotation_lines_row_count
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        new_value = count + 1000
        if new_value > len(reader):
            new_value = len(reader)

            

            self.lines_processed = True

        for i in range(count, new_value):
            lines = reader[i].split('\t')
            partner = False
            if lines[1]:
                partner = self.env['res.partner'].search([('mrm_id','=',lines[1]),('is_mrm_contact','=',False)],limit=1)
                if not partner:
                    partner = self.env['res.partner'].search([('name','=',lines[0]),('is_mrm_contact','=',False)],limit=1)

            if str(partner) == 'res.partner()':
                partner = False

            elif partner != False:
                partner = partner.id
            
            if lines[1]:
                payment = self.env['account.payment'].search([('name','=',lines[2]),('from_mrm','=',True),('partner_id.name','=',lines[0]),('state','!=','cancelled')])

            else:
                payment = self.env['account.payment'].search([('name','=',lines[2]),('from_mrm','=',True),('partner_id','=',False),('state','!=','cancelled')])

            if not payment:
                payment = self.env['account.payment'].search([('name','=',lines[2]),('from_mrm','=',True),('state','!=','cancelled')])


            invoice = self.env['account.move'].search([('name','=',lines[5])])
            if payment and not payment.invoice_allocation:
                register_payments = self.env['register.payment'].search([('payment_id','=',payment.id)])
                for register_payment in register_payments:
                    line = self.env['register.payment.line'].search([('payment_registration_id','=',register_payment.id),('move_id','=',False)])
                    if line:
                        if invoice:
                            self.env.cr.execute("update register_payment_line set move_id = " + str(invoice.id) + " where id = " + str(line.id))

            lines = self.env['register.payment.line'].search([('move_id','=',invoice.id)])
            all_register_payments = lines.mapped('payment_registration_id')
            invoices = all_register_payments.mapped('payment_line_ids').mapped('move_id')
            for inv in invoices:
                inv.button_draft()
                inv._onchange_invoice_line_ids()
                inv.action_post()


            for reg in all_register_payments:
                reg.action_register_payment()
                reg.write({'is_processed': True})



        self.write({'quotation_lines_row_count': new_value})

    def set_customer(self):
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()
        users = []
        for row in reader:
            lines = row.split('\t')
            partner = self.env['res.partner'].search([('mrm_id','=',int(lines[3])),('name','=',lines[2]),('is_mrm_contact','=',False)],limit=1)
            if not partner:
                partner = self.env['res.partner'].search([('name','=',lines[2]),('is_mrm_contact','=',False)],limit=1)

            
            if not partner:
                partner = self.env['res.partner'].create({
                    'name': lines[2],
                    'mrm_id': lines[3],
                    'company_type': 'person',
                    'from_excel': True,
                    })
            invoice = self.env['account.move'].search([('name','=',lines[0])])
            for line in invoice.line_ids:
                line.write({'partner_id': partner.id})
            invoice.write({'partner_id': partner.id})


    def fix_contact_registration(self):
        value = self.processed_rows_count
        attachment = self.env['ir.attachment'].browse(self.id)
        objFile = str(base64.b64decode(attachment.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        new_value = int(value) + 1500
        if new_value > len(reader):
            new_value = len(reader)


            self.is_processed = True

        for i in range(int(value), new_value):
            lines = reader[i].split('\t')
            vals = {}
            old_partner = False
            if lines[0]:
                old_partner = self.env['res.partner'].search([('email','=',lines[2].strip()),('mrm_id','!=',int(lines[0]))],limit=1)

                real_partner = self.env['res.partner'].search([('mrm_id','=',int(lines[0])),('is_mrm_contact','=',True)],limit=1)
                if not real_partner:
                    real_partner = self.env['res.partner'].create({
                        'mrm_id': lines[0],
                        'name': lines[1],
                        'email': lines[2],
                        'mobile': lines[3],
                        'phone': lines[4],
                        'company_type': 'person',
                        'from_excel': True,
                        'is_mrm_contact': True,

                        })

                else:
                    real_partner.write({
                        'mrm_id': lines[0],
                        'name': lines[1],
                        'mobile': lines[3],
                        'phone': lines[4],
                        'email': lines[2],
                        'company_type': 'person',
                        'from_excel': True,
                        'is_mrm_contact': True,
                        })
                
                vals['partner_id'] = real_partner.id
            
            vals['registration_number'] = lines[5]
            event = self.env['event.event'].search([('mrm_id','=',int(lines[7]))])
            vals['event_id'] = event.id



            if 'partner_id' in vals and old_partner:
                registrations = self.env['event.registration'].search([('partner_id','=',old_partner.id),('event_id','=',vals['event_id']),('registration_number','=',vals['registration_number'])])
                if registrations:
                    for registration in registrations:
                        self.env.cr.execute("update event_registration set partner_id = " + str(vals['partner_id']) + " where id = " + str(registration.id))
                        registration._onchange_partner()

        self.write({'processed_rows_count': new_value})

    def import_discount_line_lines(self):
        count = self.quotation_lines_row_count
        objFile = str(base64.b64decode(self.datas).decode('UTF-8'))
        reader = objFile.splitlines()

        new_value = count + 300
        if new_value > len(reader):
            new_value = len(reader)

            

            self.lines_processed = True

        for i in range(count, new_value):
            lines = reader[i].split('\t')
            date = False
            if lines[0]:
                invoice_name = str(lines[0])


                invoice = self.env['account.move'].search([('name','=',invoice_name)])
                if not invoice:
                    invoice_vals = {}
                    invoice_vals['name'] = invoice_name
                    invoice_vals['type'] = 'out_refund' if 'CR' in invoice_name else 'out_invoice'


                    if lines[1]:
                        date = datetime.strptime(lines[1], '%b %d, %Y')
                        invoice_vals['invoice_date_due'] = date
                        invoice_vals['date'] = date
                        invoice_vals['invoice_date'] = date

                    if lines[3]:
                        try:
                            partner = self.env['res.partner'].search([('mrm_id','=',int(lines[3])),('is_mrm_contact','=',False)],limit=1)

                        except:
                            raise UserError(lines)

                        if not partner:
                            partner = self.env['res.partner'].search([('name','=',lines[2])],limit=1)

                            if not partner:
                                partner = self.env['res.partner'].create({
                                    'name': lines[2],
                                    'mrm_id': lines[3],
                                    'company_type': 'person',
                                    'from_excel': True,
                                    })

                            else:
                                partner.write({
                                    'name': lines[2],
                                    'mrm_id': lines[3],
                                    'company_type': 'person',
                                    'from_excel': True,
                                    })

                        else:
                            partner.write({
                                'name': lines[2],
                                'mrm_id': lines[3],
                                'company_type': 'person',
                                'from_excel': True,
                                })
                    else:
                        partner = self.env['res.partner'].browse(426764)

                        invoice_vals['partner_id'] = partner.id

                    invoice_vals['other_inv_ref'] = lines[4]
                    invoice_vals['note'] = lines[23]

                    if lines[20]:
                        company_name = lines[20].strip()
                        company = self.env['res.company'].search([('name','=',company_name)],limit=1)

                        if not company:
                            company = self.env['res.company'].search([('name','ilike',company_name)],limit=1)

                        invoice_vals['company_id'] = company.id
                    
                    invoice_vals['journal_id'] = self.env['account.move'].with_context(force_company=company.id, default_type='out_invoice')._get_default_journal().id
                    currency = self.env['res.currency'].search([('name','=',lines[10])])
                    invoice_vals['currency_id'] = currency.id


                    if 'partner_id' in invoice_vals:
                        invoice = self.env['account.move'].create(invoice_vals)
                        invoice._onchange_invoice_line_ids()
                        invoice.write({'name': invoice_vals['name']})

                else:
                    invoice.button_draft()
                    if invoice.payment_ids:
                        payments = invoice.payment_ids.ids
                        self.env['register.payment'].search([('payment_id','in',payments)]).filtered(lambda p: p.is_processed == True).write({'is_processed': False})

                if lines[20]:
                    company = self.env['res.company'].search([('name','=',lines[20])],limit=1)

                    if not lines[9] and lines[12] and lines[0]:
                        if float(lines[11]) != 0:
                            invoice.button_draft()
                            if invoice.payment_ids:
                                payments = invoice.payment_ids.ids
                                self.env['register.payment'].search([('payment_id','in',payments)]).filtered(lambda p: p.is_processed == True).write({'is_processed': False})
                        

                        product = self.env['product.template'].sudo().browse(6247)

                    elif lines[9] and lines[0]:
                        product = self.env['product.template'].sudo().search([('mrm_id','=',int(lines[9]))],limit=1)
                        if not product:
                            product = self.env['product.template'].sudo().create({
                                'mrm_id': int(lines[9]),
                                'name': lines[8],
                                'type': 'product' if lines[22] == 'TRUE' else 'service'
                                })
                        else:
                            product.write({
                                'name': lines[8],
                                'type': 'product' if lines[22] == 'TRUE' else 'service'
                                })
                    
                    vals = {}
                    product_variant = product.product_variant_id
                    vals['move_id'] = invoice.id
                    vals['product_id'] = product_variant.id
                    vals['partner_id'] = invoice.partner_id.id
                    vals['product_uom_id'] = product_variant.uom_id.id
                    vals['name'] = lines[8] if lines[8] else product.name
                    vals['quantity'] = float(lines[12]) if lines[12] else 0
                    vals['discount'] = float(lines[13]) if lines[13] else 0

                    if vals['discount'] == 0 and lines[14] and float(lines[14]) > 0:
                        vals['discount'] = float(lines[14]) * 100 / float(lines[11]) if float(lines[11]) > 0 else 0 

                    price = 0
                    if lines[11]:
                        price = float(lines[11])
                        if invoice.type == 'out_refund':
                            price = -price
                    
                    vals['price_unit'] = price

                    vals['account_id'] = self.env['account.account'].sudo().search([('company_id','=', invoice.company_id.id),('code','=','400000'),('name','=','Product Sales')],limit=1).id
                    vals['exclude_from_invoice_tab'] = False


                    tax = False
                    if lines[16] and float(lines[16]) > 0:
                        tax = self.env['account.tax'].sudo().search([('mrm_id','>',0),('company_id','=',company.id),('amount','=',float(lines[16]))])
                        if not tax:
                            tax = self.env['account.tax'].sudo().search([('mrm_id','>',0),('company_id','=',company.parent_id.id),('amount','=',float(lines[16]))])

                        if not tax:
                            tax = self.env['account.tax'].sudo().search([('company_id','=',company.id),('amount','=',float(lines[16]))])

                        if not tax:
                            tax = self.env['account.tax'].sudo().search([('company_id','=',company.parent_id.id),('amount','=',float(lines[16]))])

                        if not tax:
                            tax = self.env['account.tax'].sudo().search([('company_id','=',invoice.company_id.id),('amount','=',float(lines[16]))])
                            if not tax:
                                tax_name = str(lines[16]) + '%'
                                if company:
                                    tax_name = company.name + ' - ' + str(lines[16]) + '%'

                                tax = self.env['account.tax'].create({
                                    'name': tax_name,
                                    'type_tax_use': 'sale',
                                    'company_id': company.id if company else invoice.company_id.id,
                                    'amount': float(lines[16]),
                                    'description': tax_name,
                                    })




                        vals['tax_ids'] = [(6, 0, [tax.id])] if tax else [(5, 0, 0)]

                    if lines[19]:
                        term = self.env['term'].search([('name','=',lines[19].strip())],limit=1)
                        if not term:
                            term = self.env['term'].create({
                                'name': lines[19].strip()
                                })
                        vals['term_id'] = term.id
                    
                    #create empty lines or unmigrated lines
                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                    if move_line and vals['product_id'] == 5850:
                        vals['product_id'] = 5851
                        move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                        if move_line:
                            vals['product_id'] = 5852
                            move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                            if move_line:
                                vals['product_id'] = 5853
                                move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5854
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5855
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5856
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5857
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5858
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)
                                if move_line:
                                    vals['product_id'] = 5859
                                    move_line = self.env['account.move.line'].sudo().search([('product_id','=',vals['product_id']),('move_id','=',vals['move_id']),('line_updated','=',False),('exclude_from_invoice_tab','=',False)],limit=1)



                    #move_line = self.env['account.move.line'].search([('move_id','=',vals['move_id']),('name','=',vals['name']),('price_unit','=',vals['price_unit'])])
                    #if move_line:
                        #raise UserError(vals['move_id'])
                    
                    #else:
                    move_line = self.env['account.move.line'].sudo().create(vals)

                invoice._onchange_invoice_line_ids()

        self.write({'quotation_lines_row_count': new_value})








